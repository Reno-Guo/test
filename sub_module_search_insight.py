import streamlit as st
import pandas as pd
import os
import re
from datetime import datetime
import io
import zipfile
import tempfile
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import plotly.express as px
from uuid import uuid4
from typing import Callable, List, Any, Dict

# 从主程序导入共享函数
def _read_excel_cached(file_or_path, sheet_name=0, engine=None):
    return pd.read_excel(file_or_path, sheet_name=sheet_name, engine=engine)

def unique_tmp_path(suggest_name: str, default_ext: str = ".xlsx") -> str:
    base, ext = os.path.splitext(suggest_name or f"result{default_ext}")
    ext = ext or default_ext
    return os.path.join("/tmp", f"{base}_{st.session_state.SID}_{uuid4().hex[:8]}{ext}")

def save_workbook_to_buffer(wb: Workbook) -> io.BytesIO:
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def render_download_section(
    buffer: io.BytesIO,
    file_name: str,
    mime_type: str,
    download_label: str,
    key_prefix: str,
    has_save: bool = False,
    save_func: Callable[[], None] | None = None,
    save_path: str | None = None,
):
    if has_save:
        col_d, col_s = st.columns(2)
        with col_d:
            st.download_button(
                label=download_label,
                data=buffer,
                file_name=file_name,
                mime=mime_type,
                key=f"{key_prefix}_download",
                use_container_width=True,
            )
        with col_s:
            if st.checkbox("💾 同时保存到 /tmp 目录", key=f"{key_prefix}_save"):
                if save_func:
                    save_func()
                st.info(f"📁 文件已保存到 {save_path}")
    else:
        st.download_button(
            label=download_label,
            data=buffer,
            file_name=file_name,
            mime=mime_type,
            key=f"{key_prefix}_download",
            use_container_width=True,
        )

def render_app_header(emoji_title: str, subtitle: str):
    st.markdown(f"""
    <div style="background: linear-gradient(135deg, #00a6e4 0%, #0088c2 100%); padding: 2rem; border-radius: 10px; margin-bottom: 2rem; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
        <h2 style="color: white; margin: 0; display: flex; align-items: center;">
            {emoji_title}
        </h2>
        <p style="color: rgba(255,255,255,0.9); margin-top: 0.5rem;">{subtitle}</p>
    </div>
    """, unsafe_allow_html=True)

def get_timestamp() -> str:
    return datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

def analyze_search_rows(df: pd.DataFrame, params: List[tuple]):
    punct = str.maketrans("", "", '!"#$%&\'()*+,-./:;<=>?@[\\]^_`{|}~')
    brands = df["品牌名称"].dropna().unique()
    for p, _ in params:
        df[p] = ""
    df["品牌"] = ""
    df["特性参数"] = ""
    results = []
    brand_words = []
    pb = st.progress(0)
    status = st.empty()
    for idx, row in df.iterrows():
        status.text(f"正在分析第 {idx+1}/{len(df)} 条数据...")
        sword = str(row["搜索词"]).lower()
        vol = row["搜索量"] if pd.notna(row["搜索量"]) else 0
        m_brands = []
        for b in brands:
            b_low = str(b).lower()
            if len(b_low) <= 5:
                if re.search(rf"\b{re.escape(b_low)}\b", sword):
                    m_brands.append(b_low)
            else:
                norms = [
                    b_low,
                    b_low.translate(punct),
                    b_low.replace(" ", ""),
                    b_low.translate(punct).replace(" ", ""),
                ]
                if any(n in sword for n in norms):
                    m_brands.append(b_low)
        df.at[idx, "品牌"] = ",".join(set(m_brands))
        m_params = []
        for p_name, p_vals in params:
            m_vals = [str(v).lower() for v in p_vals if str(v).lower() in sword]
            df.at[idx, p_name] = ",".join(set(m_vals))
            m_params.extend(m_vals)
        df.at[idx, "特性参数"] = ",".join(set(m_params))
        if m_brands:
            results.append("Branded KWs")
            for b in set(m_brands):
                brand_words.append({"品牌名称": b, "搜索量": vol})
        else:
            results.append("Non-Branded KWs")
        pb.progress((idx + 1) / len(df))
    status.empty()
    pb.empty()
    df["词性"] = results
    return df, results

def search_insight_app():
    render_app_header("🔍 SI - 搜索流量洞察", "分析搜索关键词，识别品牌词与非品牌词")
    st.markdown("#### 📋 步骤 1: 下载数据模板")
    tmpl = pd.DataFrame(columns=["搜索词", "搜索量", "品牌名称"])
    buf = io.BytesIO()
    tmpl.to_excel(buf, index=False)
    buf.seek(0)
    st.download_button(
        "📥 下载Excel模板",
        buf,
        "template.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_template",
        use_container_width=True,
    )
    st.divider()
    st.markdown("#### 📤 步骤 2: 上传填写好的数据文件")
    col1, col2 = st.columns([2, 1])
    with col1:
        uploaded_file = st.file_uploader("选择数据文件", type=["xlsx", "xls"], key="data_file")
    with col2:
        save_filename = st.text_input("输出文件名", "search_insight_result.xlsx", key="save_folder")
    st.divider()
    st.markdown("#### ⚙️ 步骤 3: 输入产品参数(可选)")
    col1, col2 = st.columns(2)
    with col1:
        param_names = st.text_input("参数名(用逗号分隔)", placeholder="例如: 颜色,尺寸,材质", key="param_names")
    with col2:
        param_values = st.text_area(
            "具体参数(每行一个参数组,用逗号分隔)",
            placeholder="例如:\n红,蓝,绿\n小,中,大",
            key="param_values",
            height=100,
        )
    st.divider()
    execute_btn = st.button("🚀 开始分析", key="execute_button", use_container_width=True)
    if execute_btn:
        if not uploaded_file or not save_filename:
            st.warning("⚠️ 请确保已上传数据文件并输入输出文件名")
            return
        with st.spinner("🔄 正在分析数据，请稍候..."):
            save_path = unique_tmp_path(save_filename)
            df = _read_excel_cached(uploaded_file)
            if df.empty:
                st.warning("📂 上传的文件为空，请检查数据文件")
                return
            p_params = []
            if param_names and param_values:
                names = [n.strip() for n in re.split(r"[,\uff0c]", param_names) if n.strip()]
                vals = []
                for line in param_values.split("\n"):
                    vs = [v.strip() for v in re.split(r"[,\uff0c]", line) if v.strip()]
                    if vs:
                        vals.append(vs)
                p_params = list(zip(names, vals)) if len(names) == len(vals) else []
            df, kw_types = analyze_search_rows(df, p_params)
            branded = kw_types.count("Branded KWs")
            non_branded = len(kw_types) - branded
            status = st.empty()
            prog = st.progress(0)
            status.text("正在保存到Excel...")
            prog.progress(0.5)
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            ws = wb.create_sheet("源数据")
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            prog.progress(1.0)
            status.text("保存完成")
            status.empty()
            prog.empty()
            buffer = save_workbook_to_buffer(wb)
            ts = get_timestamp()
            out_name = f"result_{ts}.xlsx"
            out_path = os.path.join("/tmp", out_name)
            st.success(f"✅ 分析完成! 品牌词: {branded} 条 | 非品牌词: {non_branded} 条")
            save_func = lambda: wb.save(out_path)
            render_download_section(
                buffer,
                out_name,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "📥 下载处理结果",
                "result",
                has_save=True,
                save_func=save_func,
                save_path=out_path,
            )
