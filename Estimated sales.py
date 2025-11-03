import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# 页面配置
st.set_page_config(
    page_title="关键词预估销量工具",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# 自定义CSS样式
st.markdown("""
<style>
    /* 主色调 */
    :root {
        --primary-color: #00a6e4;
        --primary-light: #33b8ea;
        --primary-dark: #0087b8;
        --bg-light: #f0f9fd;
        --shadow: 0 2px 8px rgba(0, 166, 228, 0.1);
    }
    
    /* 隐藏默认的Streamlit样式 */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    
    /* 主容器样式 */
    .main {
        background: linear-gradient(135deg, #f0f9fd 0%, #ffffff 100%);
    }
    
    /* 标题样式 */
    h1 {
        color: #00a6e4 !important;
        font-size: 2.5rem !important;
        font-weight: 700 !important;
        text-align: center;
        padding: 1.5rem 0;
        margin-bottom: 2rem;
        border-bottom: 3px solid #00a6e4;
        text-shadow: 2px 2px 4px rgba(0, 166, 228, 0.1);
    }
    
    /* 子标题样式 */
    h2, h3 {
        color: #0087b8 !important;
        font-weight: 600 !important;
        margin-top: 2rem !important;
    }
    
    /* 上传区域样式 */
    .uploadedFile {
        background: white !important;
        border: 2px dashed #00a6e4 !important;
        border-radius: 12px !important;
        padding: 1.5rem !important;
        margin: 1rem 0 !important;
        box-shadow: var(--shadow);
        transition: all 0.3s ease;
    }
    
    .uploadedFile:hover {
        border-color: #33b8ea !important;
        box-shadow: 0 4px 12px rgba(0, 166, 228, 0.2);
        transform: translateY(-2px);
    }
    
    /* 文件上传按钮 */
    [data-testid="stFileUploader"] {
        background: white;
        padding: 1.5rem;
        border-radius: 12px;
        box-shadow: var(--shadow);
        border: 1px solid #e0f4fc;
    }
    
    [data-testid="stFileUploader"] label {
        color: #0087b8 !important;
        font-weight: 600 !important;
        font-size: 1.1rem !important;
    }
    
    /* 按钮样式 */
    .stButton > button {
        background: linear-gradient(135deg, #00a6e4 0%, #0087b8 100%) !important;
        color: white !important;
        border: none !important;
        padding: 0.75rem 2rem !important;
        font-size: 1.1rem !important;
        font-weight: 600 !important;
        border-radius: 8px !important;
        box-shadow: 0 4px 12px rgba(0, 166, 228, 0.3) !important;
        transition: all 0.3s ease !important;
        width: 100%;
    }
    
    .stButton > button:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 6px 16px rgba(0, 166, 228, 0.4) !important;
        background: linear-gradient(135deg, #33b8ea 0%, #00a6e4 100%) !important;
    }
    
    .stButton > button:active {
        transform: translateY(0) !important;
    }
    
    /* 下载按钮特殊样式 */
    .stDownloadButton > button {
        background: linear-gradient(135deg, #28a745 0%, #20803a 100%) !important;
    }
    
    .stDownloadButton > button:hover {
        background: linear-gradient(135deg, #34ce57 0%, #28a745 100%) !important;
    }
    
    /* 信息框样式 */
    .stAlert {
        background: white !important;
        border-left: 4px solid #00a6e4 !important;
        border-radius: 8px !important;
        padding: 1rem 1.5rem !important;
        box-shadow: var(--shadow);
    }
    
    /* 表格表头 */
    [data-testid="stDataFrame"] thead tr th {
        background: linear-gradient(135deg, #00a6e4 0%, #0087b8 100%) !important;
        color: white !important;
        font-weight: 600 !important;
        padding: 0.75rem !important;
        border: none !important;
    }
    
    /* 表格行 */
    [data-testid="stDataFrame"] tbody tr:hover {
        background: #f0f9fd !important;
    }
    
    /* 卡片容器 */
    .card {
        background: white;
        padding: 1.5rem;
        border-radius: 12px;
        box-shadow: var(--shadow);
        margin: 1rem 0;
        border: 1px solid #e0f4fc;
    }
    
    /* 成功消息 */
    .success-message {
        background: linear-gradient(135deg, #d4edda 0%, #c3e6cb 100%);
        color: #155724;
        padding: 1rem;
        border-radius: 8px;
        border-left: 4px solid #28a745;
        margin: 1rem 0;
        font-weight: 500;
    }
    
    /* 步骤指示器 */
    .step-indicator {
        display: flex;
        align-items: center;
        background: white;
        padding: 1rem 1.5rem;
        border-radius: 8px;
        margin: 0.5rem 0;
        box-shadow: var(--shadow);
        border-left: 4px solid #00a6e4;
    }
    
    .step-number {
        background: #00a6e4;
        color: white;
        width: 32px;
        height: 32px;
        border-radius: 50%;
        display: flex;
        align-items: center;
        justify-content: center;
        font-weight: 700;
        margin-right: 1rem;
        flex-shrink: 0;
    }
    
    .step-text {
        color: #0087b8;
        font-weight: 500;
        font-size: 1rem;
    }
    
    /* 分隔线 */
    hr {
        border: none;
        height: 2px;
        background: linear-gradient(90deg, transparent, #00a6e4, transparent);
        margin: 2rem 0;
    }
</style>
""", unsafe_allow_html=True)

# 页面标题
st.markdown("<h1>📊 关键词预估销量工具</h1>", unsafe_allow_html=True)

st.markdown("<hr>", unsafe_allow_html=True)

# 步骤1：上传文件
st.markdown("""
<div class="step-indicator">
    <div class="step-number">1</div>
    <div class="step-text">上传第一个文件（表头在第二行，至少带有 关键词 + 搜索量排名 两列）</div>
</div>
""", unsafe_allow_html=True)

file1 = st.file_uploader(
    "选择第一个xlsx文件",
    type=["xlsx"],
    key="file1",
    help="表头在第二行，文件应包含：关键词、搜索量排名"
)

st.markdown("""
<div class="step-indicator">
    <div class="step-number">2</div>
    <div class="step-text">上传第二个文件（表头在第二行，SIF关键词转化率数据）</div>
</div>
""", unsafe_allow_html=True)

file2 = st.file_uploader(
    "选择第二个xlsx文件",
    type=["xlsx"],
    key="file2",
    help="表头在第二行，文件应包含：关键词、翻译、搜索量、点击转化率、建议竞价-推荐、建议竞价-最高、ABATop3集中度-点击"
)

if file1 and file2:
    st.markdown('<div class="success-message">✅ 文件上传成功！正在处理数据...</div>', unsafe_allow_html=True)
    
    try:
        # 读取第一个文件
        df1 = pd.read_excel(file1, skiprows=1)
        df1 = df1[['关键词', '搜索量排名']]

        # 读取第二个文件
        df2 = pd.read_excel(file2, skiprows=1)
        columns_to_keep = ['关键词', '翻译', '搜索量', '点击转化率', '建议竞价-推荐', '建议竞价-最高', 'ABATop3集中度-点击']
        df2 = df2[columns_to_keep]

        # 合并数据
        result_df = pd.merge(df2, df1, on='关键词', how='left')

        # 重新排列列顺序
        result_columns = ['关键词', '翻译', '搜索量', '点击转化率', '建议竞价-推荐', '建议竞价-最高', 'ABATop3集中度-点击', '搜索量排名']
        result_df = result_df[result_columns]

        # 添加新列
        result_df['日搜索量'] = result_df['搜索量'] / 7

        # 计算搜索量份额占比
        def calculate_share(row):
            rank = row['搜索量排名']
            bid_recommend = row['建议竞价-推荐']
            concentration = row['ABATop3集中度-点击']

            if pd.isna(rank) or pd.isna(bid_recommend) or pd.isna(concentration):
                return np.nan

            if 0 < rank <= 5000 or bid_recommend > 5:
                return 0.02
            elif 5000 < rank <= 10000:
                return 0.035
            else:
                if concentration < 0.4:
                    return 0.05
                elif 0.4 <= concentration < 0.5:
                    return 0.03
                elif 0.5 <= concentration < 0.6:
                    return 0.02
                else:
                    return 0.01

        result_df['搜索量份额占比'] = result_df.apply(calculate_share, axis=1)
        result_df['预估修正CVR'] = np.nan
        result_df['预估单量'] = np.nan

        # 更新列顺序
        final_columns = ['关键词', '翻译', '搜索量', '点击转化率', '建议竞价-推荐', '建议竞价-最高', 'ABATop3集中度-点击', '搜索量排名', '日搜索量', '搜索量份额占比', '预估修正CVR', '预估单量']
        result_df = result_df[final_columns]

        st.markdown("<hr>", unsafe_allow_html=True)
        
        # 直接使用result_df，不需要编辑
        edited_df = result_df

        # 步骤3：下载结果
        st.markdown("""
        <div class="step-indicator">
            <div class="step-number">3</div>
            <div class="step-text">下载处理结果</div>
        </div>
        """, unsafe_allow_html=True)

        # 下载按钮
        def generate_excel(df):
            wb = Workbook()
            ws = wb.active
            ws.title = "结果表"

            for r in dataframe_to_rows(df.drop(columns=['预估单量']), index=False, header=True):
                ws.append(r)

            green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            for col in range(1, 9):
                cell = ws.cell(row=1, column=col)
                cell.fill = green_fill

            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            for col in [9, 10, 11]:
                cell = ws.cell(row=1, column=col)
                cell.fill = yellow_fill

            blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            l_header = ws.cell(row=1, column=12)
            l_header.value = "预估单量"
            l_header.fill = blue_fill

            for row in range(2, len(df) + 2):
                formula = f'=I{row}*J{row}*(D{row}+K{row})'
                ws.cell(row=row, column=12).value = formula

            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output

        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            excel_file = generate_excel(edited_df)
            st.download_button(
                label="📥 下载结果Excel文件",
                data=excel_file,
                file_name="销量预估结果.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

        # 数据预览
        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown("<h3>👁️ 数据预览</h3>", unsafe_allow_html=True)
        
        preview_df = edited_df.copy()
        preview_df['预估单量'] = preview_df['日搜索量'] * preview_df['搜索量份额占比'] * (preview_df['点击转化率'] + preview_df['预估修正CVR'].fillna(0))
        
        st.dataframe(preview_df, use_container_width=True, height=400)

    except Exception as e:
        st.error(f"❌ 处理文件时出错：{str(e)}")
        st.info("请检查文件格式是否正确，确保包含所需的列名。")

else:
    st.markdown("""
    <div class="card" style="text-align: center; padding: 3rem;">
        <h3 style="color: #00a6e4;">👆 请上传两个Excel文件开始处理</h3>
        <p style="color: #666; margin-top: 1rem;">支持的文件格式：.xlsx</p>
    </div>
    """, unsafe_allow_html=True)

# 页脚
st.markdown("<hr>", unsafe_allow_html=True)
st.markdown("""
<div style="text-align: center; color: #999; padding: 2rem 0;">
    <p>© 关键词预估销量分析工具 </p>
</div>
""", unsafe_allow_html=True)
