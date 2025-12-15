import streamlit as st
import pandas as pd
import os
import re
from datetime import datetime
import io
import zipfile
import tempfile
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import plotly.express as px
from uuid import uuid4
from typing import Callable, List, Any, Dict

# === Concurrency-safe session + helpers ===
if "SID" not in st.session_state:
    st.session_state.SID = uuid4().hex[:6]

def unique_tmp_path(suggest_name: str, default_ext: str = ".xlsx") -> str:
    base, ext = os.path.splitext(suggest_name or f"result{default_ext}")
    ext = ext or default_ext
    return os.path.join("/tmp", f"{base}_{st.session_state.SID}_{uuid4().hex[:8]}{ext}")

@st.cache_data(ttl=1800, show_spinner=False)
def _read_excel_cached(file_or_path, sheet_name=0, engine=None):
    return pd.read_excel(file_or_path, sheet_name=sheet_name, engine=engine)

# App configuration
APP_CONFIG = {
    "app_title": "市场洞察小程序",
    "author": "海翼IDC团队",
    "version": "v1.2.0",
    "contact": "idc@oceanwing.com",
    "company": "Anker Oceanwing Inc."
}

# === Shared UI/render helpers ===
def render_app_header(emoji_title: str, subtitle: str):
    st.markdown(f"""
    <div style="background: linear-gradient(135deg, #00a6e4 0%, #0088c2 100%); padding: 2rem; border-radius: 10px; margin-bottom: 2rem; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
        <h2 style="color: white; margin: 0; display: flex; align-items: center;">
            {emoji_title}
        </h2>
        <p style="color: rgba(255,255,255,0.9); margin-top: 0.5rem;">{subtitle}</p>
    </div>
    """, unsafe_allow_html=True)

def get_timestamp() -> str:
    return datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

def save_df_to_buffer(df: pd.DataFrame) -> io.BytesIO:
    buffer = io.BytesIO()
    df.to_excel(buffer, index=False, engine="openpyxl")
    buffer.seek(0)
    return buffer

def save_workbook_to_buffer(wb: Workbook) -> io.BytesIO:
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def render_download_section(
    buffer: io.BytesIO,
    file_name: str,
    mime_type: str,
    download_label: str,
    key_prefix: str,
    has_save: bool = False,
    save_func: Callable[[], None] | None = None,
    save_path: str | None = None,
):
    if has_save:
        col_d, col_s = st.columns(2)
        with col_d:
            st.download_button(
                label=download_label,
                data=buffer,
                file_name=file_name,
                mime=mime_type,
                key=f"{key_prefix}_download",
                use_container_width=True,
            )
        with col_s:
            if st.checkbox("💾 同时保存到 /tmp 目录", key=f"{key_prefix}_save"):
                if save_func:
                    save_func()
                st.info(f"📁 文件已保存到 {save_path}")
    else:
        st.download_button(
            label=download_label,
            data=buffer,
            file_name=file_name,
            mime=mime_type,
            key=f"{key_prefix}_download",
            use_container_width=True,
        )

# === Shared ZIP processors ===
def read_file_merge(file_path: str) -> pd.DataFrame | None:
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        return pd.read_csv(file_path)
    engine = "openpyxl" if ext == ".xlsx" else "xlrd" if ext == ".xls" else None
    if engine:
        return _read_excel_cached(file_path, engine=engine)
    return None

def read_file_clean(file_path: str) -> pd.DataFrame | None:
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        return pd.read_csv(file_path, header=None)
    engine = "openpyxl" if ext == ".xlsx" else "xlrd" if ext == ".xls" else None
    if engine:
        return pd.read_excel(file_path, header=None, engine=engine)
    return None

def write_processed_file(df: pd.DataFrame, path: str, ext: str):
    if ext == ".csv":
        df.to_csv(path, index=False)
    else:
        df.to_excel(path, index=False, engine="openpyxl")

def process_zip_files(
    uploaded_file,
    read_cb: Callable[[str], pd.DataFrame | None],
    process_cb: Callable[[pd.DataFrame, str, str], Any],
) -> List[Any]:
    with tempfile.TemporaryDirectory() as temp_dir:
        zip_path = os.path.join(temp_dir, uploaded_file.name)
        with open(zip_path, "wb") as f:
            f.write(uploaded_file.getbuffer())
        with zipfile.ZipFile(zip_path, "r") as z:
            z.extractall(temp_dir)
        files = [f for f in os.listdir(temp_dir) if f.lower().endswith((".xlsx", ".xls", ".csv"))]
        if not files:
            st.warning("📂 压缩文件中未找到任何 Excel 或 CSV 文件")
            return []
        results = []
        pb = st.progress(0)
        status = st.empty()
        for i, f in enumerate(files):
            status.text(f"正在处理: {f} ({i+1}/{len(files)})")
            fp = os.path.join(temp_dir, f)
            try:
                df = read_cb(fp)
                if df is None:
                    raise ValueError("不支持的文件格式")
                results.append(process_cb(df, f, temp_dir))
            except Exception as e:
                st.error(f"❌ 处理文件 {f} 失败: {e}")
            pb.progress((i + 1) / len(files))
        status.empty()
        pb.empty()
        return results

# 处理价格列
def process_price_columns(df):
    df = df.copy()
    price_pattern = re.compile(r'\$(\d+\.\d+)(?:\s*-\s*\$\d+\.\d+)?')
    def extract_price(price_str):
        if not isinstance(price_str, str):
            return price_str
        price_str = price_str.replace(',', '')
        match = price_pattern.match(price_str)
        return float(match.group(1)) if match else float(price_str.replace('$', ''))
    price_columns = [col for col in df.columns if '售价' in col]
    for column in price_columns:
        df[column] = df[column].apply(extract_price)
    return df

# 合并数据表格功能
def merge_data_app():
    render_app_header("📊 MI/SI - 合并数据表格", "将多个Excel文件合并为一个统一的数据表格")
    col1, col2 = st.columns([2, 1])
    with col1:
        uploaded_file = st.file_uploader(
            "选择一个 .zip 文件(包含需要合并的 Excel 文件)",
            type=["zip"],
            accept_multiple_files=False,
            key="merge_files",
            help="支持包含.xlsx、.xls、.csv格式的ZIP压缩包",
        )
    with col2:
        save_filename = st.text_input(
            "输出文件名",
            value="merged_output.xlsx",
            key="merge_save",
            help="请输入合并后的文件名",
        )
    st.divider()
    execute_btn = st.button("🚀 开始合并", key="merge_button", use_container_width=True)
    if execute_btn:
        if not uploaded_file or not save_filename:
            st.warning("⚠️ 请确保已选择 .zip 文件并输入文件名")
            return
        with st.spinner("🔄 正在处理文件，请稍候..."):
            save_path = unique_tmp_path(save_filename)
            def cb_merge(df, fname, _):
                df["时间"] = os.path.splitext(fname)[0]
                return process_price_columns(df)
            df_list = process_zip_files(uploaded_file, read_file_merge, cb_merge)
            if not df_list:
                return
            status = st.empty()
            prog = st.progress(0)
            status.text("正在合并数据...")
            merged_df = pd.concat(df_list, ignore_index=True)
            merged_df = merged_df.loc[:, ~merged_df.columns.duplicated()]
            prog.progress(1.0)
            status.text("合并完成")
            status.empty()
            prog.empty()
            buffer = save_df_to_buffer(merged_df)
            st.success(f"✅ 成功合并 {len(df_list)} 个文件，共 {len(merged_df)} 行数据")
            save_func = lambda: merged_df.to_excel(save_path, index=False, engine="openpyxl")
            render_download_section(
                buffer,
                os.path.basename(save_filename),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "📥 下载合并后的文件",
                "merged",
                has_save=True,
                save_func=save_func,
                save_path=save_path,
            )

# 搜索流量洞察（源数据）
def analyze_search_rows(df: pd.DataFrame, params: List[tuple]):
    punct = str.maketrans("", "", '!"#$%&\'()*+,-./:;<=>?@[\\]^_`{|}~')
    brands = df["品牌名称"].dropna().unique()
    for p, _ in params:
        df[p] = ""
    df["品牌"] = ""
    df["特性参数"] = ""
    results = []
    brand_words = []
    pb = st.progress(0)
    status = st.empty()
    for idx, row in df.iterrows():
        status.text(f"正在分析第 {idx+1}/{len(df)} 条数据...")
        sword = str(row["搜索词"]).lower()
        vol = row["搜索量"] if pd.notna(row["搜索量"]) else 0
        m_brands = []
        for b in brands:
            b_low = str(b).lower()
            if len(b_low) <= 5:
                if re.search(rf"\b{re.escape(b_low)}\b", sword):
                    m_brands.append(b_low)
            else:
                norms = [
                    b_low,
                    b_low.translate(punct),
                    b_low.replace(" ", ""),
                    b_low.translate(punct).replace(" ", ""),
                ]
                if any(n in sword for n in norms):
                    m_brands.append(b_low)
        df.at[idx, "品牌"] = ",".join(set(m_brands))
        m_params = []
        for p_name, p_vals in params:
            m_vals = [str(v).lower() for v in p_vals if str(v).lower() in sword]
            df.at[idx, p_name] = ",".join(set(m_vals))
            m_params.extend(m_vals)
        df.at[idx, "特性参数"] = ",".join(set(m_params))
        if m_brands:
            results.append("Branded KWs")
            for b in set(m_brands):
                brand_words.append({"品牌名称": b, "搜索量": vol})
        else:
            results.append("Non-Branded KWs")
        pb.progress((idx + 1) / len(df))
    status.empty()
    pb.empty()
    df["词性"] = results
    return df, results

def search_insight_app():
    render_app_header("🔍 SI - 搜索流量洞察", "分析搜索关键词，识别品牌词与非品牌词")
    st.markdown("#### 📋 步骤 1: 下载数据模板")
    tmpl = pd.DataFrame(columns=["搜索词", "搜索量", "品牌名称"])
    buf = io.BytesIO()
    tmpl.to_excel(buf, index=False)
    buf.seek(0)
    st.download_button(
        "📥 下载Excel模板",
        buf,
        "template.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_template",
        use_container_width=True,
    )
    st.divider()
    st.markdown("#### 📤 步骤 2: 上传填写好的数据文件")
    col1, col2 = st.columns([2, 1])
    with col1:
        uploaded_file = st.file_uploader("选择数据文件", type=["xlsx", "xls"], key="data_file")
    with col2:
        save_filename = st.text_input("输出文件名", "search_insight_result.xlsx", key="save_folder")
    st.divider()
    st.markdown("#### ⚙️ 步骤 3: 输入产品参数(可选)")
    col1, col2 = st.columns(2)
    with col1:
        param_names = st.text_input("参数名(用逗号分隔)", placeholder="例如: 颜色,尺寸,材质", key="param_names")
    with col2:
        param_values = st.text_area(
            "具体参数(每行一个参数组,用逗号分隔)",
            placeholder="例如:\n红,蓝,绿\n小,中,大",
            key="param_values",
            height=100,
        )
    st.divider()
    execute_btn = st.button("🚀 开始分析", key="execute_button", use_container_width=True)
    if execute_btn:
        if not uploaded_file or not save_filename:
            st.warning("⚠️ 请确保已上传数据文件并输入输出文件名")
            return
        with st.spinner("🔄 正在分析数据，请稍候..."):
            save_path = unique_tmp_path(save_filename)
            df = _read_excel_cached(uploaded_file)
            if df.empty:
                st.warning("📂 上传的文件为空，请检查数据文件")
                return
            p_params = []
            if param_names and param_values:
                names = [n.strip() for n in re.split(r"[,\uff0c]", param_names) if n.strip()]
                vals = []
                for line in param_values.split("\n"):
                    vs = [v.strip() for v in re.split(r"[,\uff0c]", line) if v.strip()]
                    if vs:
                        vals.append(vs)
                p_params = list(zip(names, vals)) if len(names) == len(vals) else []
            df, kw_types = analyze_search_rows(df, p_params)
            branded = kw_types.count("Branded KWs")
            non_branded = len(kw_types) - branded
            status = st.empty()
            prog = st.progress(0)
            status.text("正在保存到Excel...")
            prog.progress(0.5)
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            ws = wb.create_sheet("源数据")
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            prog.progress(1.0)
            status.text("保存完成")
            status.empty()
            prog.empty()
            buffer = save_workbook_to_buffer(wb)
            ts = get_timestamp()
            out_name = f"result_{ts}.xlsx"
            out_path = os.path.join("/tmp", out_name)
            st.success(f"✅ 分析完成! 品牌词: {branded} 条 | 非品牌词: {non_branded} 条")
            save_func = lambda: wb.save(out_path)
            render_download_section(
                buffer,
                out_name,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "📥 下载处理结果",
                "result",
                has_save=True,
                save_func=save_func,
                save_path=out_path,
            )

# 可视化共享
def aggregate_top_n(df, value_col, name_col, top_n=10):
    df = df.copy()
    df[name_col] = df[name_col].astype(str)
    df = df.sort_values(by=value_col, ascending=False).reset_index(drop=True)
    if len(df) > top_n:
        top_df = df.iloc[:top_n]
        others = df.iloc[top_n:][value_col].sum()
        others_row = pd.DataFrame([{name_col: "Others", value_col: others}])
        df = pd.concat([top_df[[name_col, value_col]], others_row], ignore_index=True)
    return df[[name_col, value_col]]

def pie_chart(df, value_col, name_col, title):
    df = df.copy()
    df[name_col] = df[name_col].astype(str)
    df = df.sort_values(by=value_col, ascending=False).reset_index(drop=True)
    if "Others" in df[name_col].values:
        order = [n for n in df[name_col] if n != "Others"] + ["Others"]
        df[name_col] = pd.Categorical(df[name_col], categories=order, ordered=True)
    palette = [
        "#4C8EDA", "#FFA14E", "#F25C5C", "#6BD0C1", "#58C27D", "#F7C948",
        "#B685D6", "#FF90B3", "#BC8D6E", "#C9C9C9", "#81D3EB",
    ]
    fig = px.pie(
        df,
        values=value_col,
        names=name_col,
        title=title,
        color_discrete_sequence=palette,
    )
    fig.update_traces(textinfo="label+percent", sort=False)
    fig.update_layout(
        height=900,
        legend=dict(orientation="v", x=0.8, y=0.5, font=dict(size=16)),
        margin=dict(l=20, r=150, t=50, b=50),
        font=dict(size=16),
    )
    st.plotly_chart(fig, use_container_width=True)

def search_insight_viz_app():
    render_app_header("📈 SI - 搜索流量洞察: 聚合和可视化", "生成多维度数据分析报表和可视化图表")
    col1, col2 = st.columns([2, 1])
    with col1:
        uploaded_file = st.file_uploader(
            "选择包含源数据的 Excel 文件(完成检查确认无误)",
            type=["xlsx", "xls"],
            key="viz_data_file",
        )
    with col2:
        save_filename = st.text_input("输出文件名", "viz_result.xlsx", key="viz_save_folder")
    st.divider()
    execute_btn = st.button("🚀 开始可视化", key="viz_execute_button", use_container_width=True)
    if execute_btn:
        if not uploaded_file or not save_filename:
            st.warning("⚠️ 请确保已上传数据文件并输入输出文件名")
            return
        with st.spinner("🔄 正在生成可视化报表，请稍候..."):
            save_path = unique_tmp_path(save_filename)
            df = _read_excel_cached(uploaded_file, sheet_name="源数据")
            if df.empty:
                st.warning("📂 上传的文件为空或不包含'源数据'工作表，请检查数据文件")
                return
            # Brand aggregation
            brand_words = []
            b_status = st.empty()
            b_prog = st.progress(0)
            b_status.text("正在处理品牌词...")
            step = max(1, len(df) // 10)
            for idx, row in df.iterrows():
                vol = row["搜索量"] if pd.notna(row["搜索量"]) else 0
                brands = [b.strip() for b in str(row["品牌"]).split(",") if b.strip()]
                for b in brands:
                    brand_words.append({"品牌名称": b, "搜索量": vol})
                if (idx + 1) % step == 0 or idx == len(df) - 1:
                    b_prog.progress((idx + 1) / len(df))
            brand_df = pd.DataFrame()
            if brand_words:
                brand_df = pd.DataFrame(brand_words).groupby("品牌名称", as_index=False)["搜索量"].sum()
                brand_df = aggregate_top_n(brand_df, "搜索量", "品牌名称")
            b_status.text("品牌词处理完成")
            b_prog.empty()
            b_status.empty()
            # Param aggregation (single pass)
            excluded = {"搜索词", "搜索量", "品牌名称", "品牌", "特性参数", "词性"}
            param_cols = [c for c in df.columns if c not in excluded]
            param_heats: Dict[str, List[Dict]] = {c: [] for c in param_cols}
            p_status = st.empty()
            p_prog = st.progress(0)
            p_status.text("正在处理参数...")
            for idx, row in df.iterrows():
                vol = row["搜索量"] if pd.notna(row["搜索量"]) else 0
                for c in param_cols:
                    val = str(row[c]) if pd.notna(row[c]) else ""
                    for v in [v.strip() for v in val.split(",") if v.strip()]:
                        param_heats[c].append({"参数值": v, "搜索量": vol})
                if (idx + 1) % step == 0 or idx == len(df) - 1:
                    p_prog.progress((idx + 1) / len(df))
            p_status.text("参数处理完成")
            p_prog.empty()
            p_status.empty()
            # Traffic structure
            traffic_df = df[["词性", "搜索量"]].groupby("词性", as_index=False)["搜索量"].sum()
            traffic_df = aggregate_top_n(traffic_df, "搜索量", "词性")
            # Workbook
            s_status = st.empty()
            s_prog = st.progress(0)
            s_status.text("正在生成Excel工作簿...")
            s_prog.progress(0.3)
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            ws = wb.create_sheet("源数据")
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            s_prog.progress(0.6)
            if not brand_df.empty:
                ws = wb.create_sheet("品牌词拆解")
                for r in dataframe_to_rows(brand_df, index=False, header=True):
                    ws.append(r)
            s_prog.progress(0.7)
            param_dfs: Dict[str, pd.DataFrame] = {}
            active_params = [c for c in param_cols if param_heats[c]]
            for i, c in enumerate(active_params):
                heats = param_heats[c]
                if heats:
                    pdf = pd.DataFrame(heats).groupby("参数值", as_index=False)["搜索量"].sum()
                    pdf = aggregate_top_n(pdf, "搜索量", "参数值")
                    param_dfs[c] = pdf
                    clean = re.sub(r"[\/*?[\]]", "", c)[:31]
                    ws = wb.create_sheet(f"{clean}拆解")
                    for r in dataframe_to_rows(pdf, index=False, header=True):
                        ws.append(r)
                s_prog.progress(0.7 + 0.3 * (i + 1) / max(1, len(active_params)))
            if not traffic_df.empty:
                ws = wb.create_sheet("品类流量结构")
                for r in dataframe_to_rows(traffic_df, index=False, header=True):
                    ws.append(r)
            s_prog.progress(1.0)
            s_status.text("工作簿生成完成")
            s_status.empty()
            s_prog.empty()
            buffer = save_workbook_to_buffer(wb)
            st.success("✅ 数据处理完成，正在生成可视化图表...")
            st.markdown("### 📊 数据可视化")
            if not brand_df.empty:
                pie_chart(brand_df, "搜索量", "品牌名称", "品牌词拆解")
            for c in param_cols:
                if c in param_dfs:
                    pie_chart(param_dfs[c], "搜索量", "参数值", f"{c} 参数搜索量分布")
            if not traffic_df.empty:
                pie_chart(traffic_df, "搜索量", "词性", "流量结构")
            st.divider()
            ts = get_timestamp()
            out_name = f"viz_result_{ts}.xlsx"
            out_path = os.path.join("/tmp", out_name)
            save_func = lambda: wb.save(out_path)
            render_download_section(
                buffer,
                out_name,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "📥 下载完整报表",
                "viz",
                has_save=True,
                save_func=save_func,
                save_path=out_path,
            )

# 数据清理
def data_clean_app():
    render_app_header("🧹 DC - 数据清理: 删除第一行", "批量删除Excel/CSV文件的第一行数据并重新打包")
    col1, col2 = st.columns([2, 1])
    with col1:
        uploaded_file = st.file_uploader(
            "选择一个 .zip 文件(包含 XLSX 或 CSV 文件)",
            type=["zip"],
            key="clean_files",
        )
    with col2:
        output_filename = st.text_input("输出文件名", "cleaned_files.zip", key="clean_save")
    st.divider()
    execute_btn = st.button("🚀 开始清理", key="clean_button", use_container_width=True)
    if execute_btn:
        if not uploaded_file or not output_filename:
            st.warning("⚠️ 请确保已选择 .zip 文件并输入输出文件名")
            return
        with st.spinner("🔄 正在清理文件，请稍候..."):
            def cb_clean(df, fname, tdir):
                df = df.iloc[1:].reset_index(drop=True)
                out_path = os.path.join(tdir, f"cleaned_{fname}")
                ext = os.path.splitext(fname)[1].lower()
                write_processed_file(df, out_path, ext)
                return out_path
            processed = process_zip_files(uploaded_file, read_file_clean, cb_clean)
            if not processed:
                return
            status = st.empty()
            prog = st.progress(0)
            status.text("正在打包ZIP文件...")
            buffer = io.BytesIO()
            with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED, compresslevel=9) as nz:
                nz.setpassword(None)
                for i, p in enumerate(processed):
                    original_name = os.path.basename(p).replace("cleaned_", "")
                    nz.write(p, original_name, compress_type=zipfile.ZIP_DEFLATED)
                    prog.progress((i + 1) / len(processed))
            buffer.seek(0)
            status.text("打包完成")
            status.empty()
            prog.empty()
            st.success(f"✅ 成功清理 {len(processed)} 个文件")
            render_download_section(
                buffer,
                output_filename,
                "application/zip",
                "📥 下载清理后的 ZIP 文件",
                "cleaned",
                has_save=False,
            )

# 剂型打标工具（新整合的功能）
class PackFormLabeler:
    def __init__(self):
        """初始化剂型分类和正则表达式模式"""
        self.pack_forms = {
            'Capsule': [
                # 英文
                r'\bcapsule\b', r'\bcapsules\b', r'\bcap\b', r'\bcaps\b',
                r'\bgelcap\b', r'\bgelcaps\b', 
                # 中文
                r'\b胶囊\b', r'\b软胶囊\b', r'\b硬胶囊\b', r'\b肠溶胶囊\b',
                r'\b缓释胶囊\b', r'\b控释胶囊\b'
            ],
            'Tablet': [
                # 英文
                r'\btablet\b',r'\bcaplet\b', r'\btablets\b', r'\btab\b', r'\btabs\b',
                r'\bchewable\b',    r'\bchewables\b', r'\bsublingual\b', r'\benteric\b', r'\bCaplets\b', 
                # 中文
                r'\b片剂\b', r'\b片\b', r'\b咀嚼片\b', r'\b含片\b',
                r'\b舌下片\b', r'\b肠溶片\b', r'\b缓释片\b', r'\b控释片\b'
            ],
            'Powder': [
                # 英文
                r'\bpowder\b', r'\bpowders\b', r'\bpwd\b', r'\bgranule\b',
                r'\bgranules\b', r'\bdrink\b', r'\bdrinks\b',r'\bCrystal\b',
                # 中文
                r'\b粉剂\b', r'\b粉末\b', r'\b冲剂\b', r'\b散剂\b',
                r'\b颗粒剂\b', r'\b冲饮\b', r'\b饮品\b'
            ],
            'Gummy': [
                # 英文
                r'\bgummy\b', r'\bgummies\b',r'\bGummy\b', r'\bGummies\b',
                r'\bcandy\b', r'\bcandies\b', r'\bjelly\b', r'\bjellies\b',
                # 中文
                r'软糖', r'咀嚼糖', r'果冻', r'糖果',
                r'口香糖', r'咀嚼片'
            ],
            'Drop': [
                # 英文
                r'\bdrop\b', r'\bdrops\b', r'\btincture\b', r'\btinctures\b',
                r'\bessence\b', r'\bessences\b', r'\bFL OZs\b',
                r'\bliquid\s*drop\b', r'\bliquid\s*drops\b',
                # 中文
                r'滴剂', r'滴液', r'酊剂', r'精华',
                r'精华液', r'液体滴剂', r'液体滴液'
            ],
            'Softgel': [
                # 英文
                r'\bsoftgel\b', r'\bsoftgels\b', r'\bsoft\s*gel\b',
                r'\bgel\b', r'\bgels\b', r'\bgelatin\b',
                # 中文
                r'软胶囊', r'软胶', r'明胶'
            ],
            'Liquid': [
                # 英文
                r'\bliquid\b', r'\bliquids\b', r'\bsyrup\b', r'\bsyrups\b',
                r'\bsuspension\b', r'\bsuspensions\b', r'\belixir\b',
                r'\bsolution\b', r'\bsolutions\b', r'\bemulsion\b',
                # 中文
                r'液体', r'口服液', r'糖浆', r'混悬液',
                r'溶液', r'乳剂', r'水剂'
            ],
            'Cream': [
                # 英文
                r'\bcream\b', r'\bcreams\b', r'\bointment\b', r'\bointments\b',
                # 中文
                r'乳膏', r'霜剂', r'软膏', r'膏剂'
            ],
            'Spray': [
                # 英文
                r'\bspray\b', r'\bsprays\b', r'\binhaler\b', r'\binhalers\b',
                # 中文
                r'喷雾', r'喷剂', r'吸入器', r'吸入剂'
            ],
            'Lotion': [
                # 英文
                r'\blotion\b', r'\blotions\b',
                # 中文
                r'乳液', r'洗剂'
            ],
            'Patch': [
                # 英文
                r'\bpatch\b', r'\bpatches\b',
                # 中文
                r'贴剂', r'贴片', r'贴膏'
            ],
            'Suppository': [
                # 英文
                r'\bsuppository\b', r'\bsuppositories\b',
                # 中文
                r'栓剂', r'坐药'
            ],
            'Oil': [
                # 英文
                r'\boil\b', r'\boils\b', r'\boils\b',
                r'\bessential\s*oil\b', r'\bessential\s*oils\b',
                r'\bfish\s*oil\b', r'\bomega\s*oil\b',
                r'\bcarrier\s*oil\b', r'\bcarrier\s*oils\b',
                # 中文
                r'油', r'精油', r'鱼油', r'植物油', r'橄榄油',
                r'椰子油', r'亚麻籽油', r'月见草油'
            ]
        }
        
        # 标准化映射表 
        self.standardization_map = {
    # ========================================
    # Capsule 相关
    # ========================================
    'capsule': 'Capsule', 'capsules': 'Capsule',
    'cap': 'Capsule', 'caps': 'Capsule', 'capsu': 'Capsule',
    'gelcaps': 'Capsule', 'gelcap': 'Capsule',
    # 首字母大写
    'Capsule': 'Capsule', 'Capsules': 'Capsule','VegCap': 'Capsule',
    'Cap': 'Capsule', 'Caps': 'Capsule', 'Capsu': 'Capsule',
    'Gelcaps': 'Capsule', 'Gelcap': 'Capsule',
    # 全大写
    'CAPSULE': 'Capsule', 'CAPSULES': 'Capsule',
    'CAP': 'Capsule', 'CAPS': 'Capsule', 'CAPSU': 'Capsule',
    'GELCAPS': 'Capsule', 'GELCAP': 'Capsule',

    # ========================================
    # Tablet 相关（包含 caplet）
    # ========================================
    'tablet': 'Tablet', 'tablets': 'Tablet',
    'tab': 'Tablet', 'tabs': 'Tablet',
    'caplet': 'Tablet', 'caplets': 'Tablet',  # ✅ 正确归类到 Tablet
    'chewable': 'Tablet', 'chewables': 'Tablet',
    'chew': 'Tablet', 'chews': 'Tablet',
    'sublingual': 'Tablet', 'enteric': 'Tablet',
    # 首字母大写
    'Tablet': 'Tablet', 'Tablets': 'Tablet',
    'Tab': 'Tablet', 'Tabs': 'Tablet',
    'Caplet': 'Tablet', 'Caplets': 'Tablet',  # ✅ 首字母大写也归为 Tablet
    'Chewable': 'Tablet', 'Chewables': 'Tablet',
    'Chew': 'Tablet', 'Chews': 'Tablet',
    'Sublingual': 'Tablet', 'Enteric': 'Tablet',
    # 全大写
    'TABLET': 'Tablet', 'TABLETS': 'Tablet',
    'TAB': 'Tablet', 'TABS': 'Tablet',
    'CAPLET': 'Tablet', 'CAPLETS': 'Tablet',  # ✅ 全大写也正确映射
    'CHEWABLE': 'Tablet', 'CHEWABLES': 'Tablet',
    'CHEW': 'Tablet', 'CHEWS': 'Tablet',
    'SUBLINGUAL': 'Tablet', 'ENTERIC': 'Tablet',

    # ========================================
    # Powder 相关
    # ========================================
    'powder': 'Powder', 'powders': 'Powder','Powdered': 'Powder',
    'granule': 'Powder', 'granules': 'Powder',
    'Crystals': 'Powder','Crystal': 'Powder','crystal': 'Powder','crystals': 'Powder',
    'pwd': 'Powder',
    'Powder': 'Powder', 'Powders': 'Powder',
    'Granule': 'Powder', 'Granules': 'Powder',
    'Pwd': 'Powder',
    'POWDER': 'Powder', 'POWDERS': 'Powder',
    'GRANULE': 'Powder', 'GRANULES': 'Powder',
    'PWD': 'Powder',

    # ========================================
    # Gummy 相关
    # ========================================
    'gummy': 'Gummy', 'gummies': 'Gummy',
    'jelly': 'Gummy', 'jellies': 'Gummy',
    'gumm': 'Gummy',
    'Gummy': 'Gummy', 'Gummies': 'Gummy',
    'Jelly': 'Gummy', 'Jellies': 'Gummy',
    'Gumm': 'Gummy',
    'GUMMY': 'Gummy', 'GUMMIES': 'Gummy',
    'JELLY': 'Gummy', 'JELLIES': 'Gummy',
    'GUMM': 'Gummy',

    # ========================================
    # Drop 相关
    # ========================================
    'drop': 'Drop', 'drops': 'Drop',
    'tincture': 'Drop', 'tinctures': 'Drop',
    'fl oz': 'Drop', 'fl. oz.': 'Drop',
    'Drop': 'Drop', 'Drops': 'Drop',
    'Tincture': 'Drop', 'Tinctures': 'Drop',
    'Fl Oz': 'Drop', 'Fl. Oz.': 'Drop',
    'DROP': 'Drop', 'DROPS': 'Drop',
    'TINCTURE': 'Drop', 'TINCTURES': 'Drop',
    'FL OZ': 'Drop', 'FL. OZ.': 'Drop',

    # ========================================
    # Softgel 相关
    # ========================================
    'softgel': 'Softgel', 'softgels': 'Softgel','sof': 'Softgel',
    'gel': 'Softgel', 'gels': 'Softgel',
    'Softgel': 'Softgel', 'Softgels': 'Softgel',
    'Gel': 'Softgel', 'Gels': 'Softgel',
    'SOFTGEL': 'Softgel', 'SOFTGELS': 'Softgel',
    'GEL': 'Softgel', 'GELS': 'Softgel',

    # ========================================
    # Liquid 相关
    # ========================================
    'liquid': 'Liquid', 'liquids': 'Liquid',
    'syrup': 'Liquid', 'syrups': 'Liquid',
    'solution': 'Liquid', 'solutions': 'Liquid',
    'suspension': 'Liquid', 'suspensions': 'Liquid',
    'Liquid': 'Liquid', 'Liquids': 'Liquid',
    'Syrup': 'Liquid', 'Syrups': 'Liquid',
    'Solution': 'Liquid', 'Solutions': 'Liquid',
    'Suspension': 'Liquid', 'Suspensions': 'Liquid',
    'LIQUID': 'Liquid', 'LIQUIDS': 'Liquid',
    'SYRUP': 'Liquid', 'SYRUPS': 'Liquid',
    'SOLUTION': 'Liquid', 'SOLUTIONS': 'Liquid',
    'SUSPENSION': 'Liquid', 'SUSPENSIONS': 'Liquid',

    # ========================================
    # Cream 相关
    # ========================================
    'cream': 'Cream', 'creams': 'Cream',
    'ointment': 'Cream', 'ointments': 'Cream',
    'Cream': 'Cream', 'Creams': 'Cream',
    'Ointment': 'Cream', 'Ointments': 'Cream',
    'CREAM': 'Cream', 'CREAMS': 'Cream',
    'OINTMENT': 'Cream', 'OINTMENTS': 'Cream',

    # ========================================
    # Spray 相关
    # ========================================
    'spray': 'Spray', 'sprays': 'Spray',
    'inhaler': 'Spray', 'inhalers': 'Spray',
    'Spray': 'Spray', 'Sprays': 'Spray',
    'Inhaler': 'Spray', 'Inhalers': 'Spray',
    'SPRAY': 'Spray', 'SPRAYS': 'Spray',
    'INHALER': 'Spray', 'INHALERS': 'Spray',

    # ========================================
    # Lotion 相关
    # ========================================
    'lotion': 'Lotion', 'lotions': 'Lotion',
    'Lotion': 'Lotion', 'Lotions': 'Lotion',
    'LOTION': 'Lotion', 'LOTIONS': 'Lotion',

    # ========================================
    # Patch 相关
    # ========================================
    'patch': 'Patch', 'patches': 'Patch',
    'Patch': 'Patch', 'Patches': 'Patch',
    'PATCH': 'Patch', 'PATCHES': 'Patch',

    # ========================================
    # Suppository 相关
    # ========================================
    'suppository': 'Suppository', 'suppositories': 'Suppository',
    'Suppository': 'Suppository', 'Suppositories': 'Suppository',
    'SUPPOSITORY': 'Suppository', 'SUPPOSITORIES': 'Suppository',

    # ========================================
    # Oil 相关
    # ========================================
    'oil': 'Oil', 'oils': 'Oil',
    'essential oil': 'Oil', 'essential oils': 'Oil',
    'fish oil': 'Oil', 'omega oil': 'Oil',
    'carrier oil': 'Oil', 'carrier oils': 'Oil',
    'Oil': 'Oil', 'Oils': 'Oil',
    'Carrier Oil': 'Oil', 'Carrier Oils': 'Oil',
    'OIL': 'Oil', 'OILS': 'Oil',
    'CARRIER OIL': 'Oil', 'CARRIER OILS': 'Oil',

    # ========================================
    # Others 相关
    # ========================================
    'bag': 'Others', 'bags': 'Others','Tea bags': 'Others',
    'teabag': 'Others', 'teabags': 'Others',
    'strip': 'Others', 'strips': 'Others',
    'stick': 'Others', 'sticks': 'Others',
    'other': 'Others', 'others': 'Others',
    'strippy': 'Others',
    # 首字母大写
    'Bag': 'Others', 'Bags': 'Others',
    'Teabag': 'Others', 'Teabags': 'Others',
    'Strip': 'Others', 'Strips': 'Others',
    'Stick': 'Others', 'Sticks': 'Others',
    'Other': 'Others', 'Others': 'Others',
    'Strippy': 'Others',
    # 全大写
    'BAG': 'Others', 'BAGS': 'Others',
    'TEABAG': 'Others', 'TEABAGS': 'Others',
    'STRIP': 'Others', 'STRIPS': 'Others',
    'STICK': 'Others', 'STICKS': 'Others',
    'OTHER': 'Others', 'OTHERS': 'Others',
    'STRIPPY': 'Others',
    }
    
    def detect_others_forms(self, product_text):
        """
        检测Others类剂型
        
        Args:
            product_text (str): 产品描述文本
            
        Returns:
            list: 检测到的Others类剂型列表
        """
        if pd.isna(product_text) or not isinstance(product_text, str):
            return []
        
        others_patterns = {
            'Injection': [r'\binjection\b', r'\binjections\b', r'注射剂', r'针剂'],
            'Nasal': [r'\bnasal\b', r'鼻用', r'鼻腔'],
            'Topical': [r'\btopical\b', r'外用', r'局部'],
            'External': [r'\bexternal\b', r'外用', r'外部'],
            'Bag': [r'\bbag\b', r'\bbags\b', r'袋装', r'包装'],
            'Teabag': [r'\bteabag\b', r'\bteabags\b', r'茶包', r'袋泡茶'],
            'Strip': [r'\bstrip\b', r'\bstrips\b', r'条装', r'条剂'],
            'Stick': [r'\bstick\b', r'\bsticks\b', r'棒状', r'棒剂']
        }
        
        detected_others = []
        text_lower = product_text.lower()
        
        for form, patterns in others_patterns.items():
            for pattern in patterns:
                if re.search(pattern, text_lower, re.IGNORECASE):
                    detected_others.append(form)
                    break
        
        return detected_others

    def standardize_pack_form(self, pack_form):
        """
        标准化剂型名称
        
        Args:
            pack_form (str): 原始剂型名称
            
        Returns:
            str: 标准化后的剂型名称
        """
        if pd.isna(pack_form) or pack_form == '':
            return pack_form
        
        # 转换为字符串
        pack_form_str = str(pack_form).strip()
        
        # 检查是否已经在标准映射表中
        if pack_form_str in self.standardization_map:
            return self.standardization_map[pack_form_str]
        
        # 检查是否匹配正则表达式模式
        for standard_form, patterns in self.pack_forms.items():
            for pattern in patterns:
                if re.search(pattern, pack_form_str, re.IGNORECASE):
                    return standard_form
        
        # 如果没有匹配到，返回原值
        return pack_form_str
    
    def detect_pack_form(self, product_text):
        """
        从产品描述中检测剂型
        
        Args:
            product_text (str): 产品描述文本
            
        Returns:
            tuple: (检测到的剂型列表, 匹配的文本列表)
        """
        if pd.isna(product_text) or not isinstance(product_text, str):
            return [], []
        
        detected_forms = []
        matched_texts = []
        
        # 转换为小写进行匹配
        text_lower = product_text.lower()
        
        # 检查主要剂型
        for form, patterns in self.pack_forms.items():
            for pattern in patterns:
                matches = re.findall(pattern, text_lower)
                if matches:
                    detected_forms.append(form)
                    matched_texts.extend(matches)
        
        # 检查Others类剂型
        others_forms = self.detect_others_forms(product_text)
        if others_forms:
            detected_forms.append('Others')
            matched_texts.extend(others_forms)
        
        return detected_forms, matched_texts
    
    def classify_pack_form(self, detected_forms):
        """
        根据检测到的剂型进行分类
        
        Args:
            detected_forms (list): 检测到的剂型列表
            
        Returns:
            str: 分类结果
        """
        if not detected_forms:
            return 'Others'
        
        # 去重
        unique_forms = list(set(detected_forms))
        
        # 特殊处理：如果同时检测到Liquid和Drop，优先归类为Drop
        if 'Liquid' in unique_forms and 'Drop' in unique_forms:
            return 'Drop'
        
        if len(unique_forms) == 1:
            return unique_forms[0]
        elif len(unique_forms) > 1:
            return 'Bundle'
        else:
            return 'Others'
    
    def process_dataframe(self, df):
        """
        处理DataFrame，对Pack form列进行智能打标和标准化
        
        Args:
            df (pd.DataFrame): 包含'Pack form'和'Product'列的DataFrame
            
        Returns:
            pd.DataFrame: 处理后的DataFrame
        """
        # 复制DataFrame避免修改原始数据
        df_processed = df.copy()
        
        # 添加新列
        df_processed['Matched_Pack_Form'] = ''
        df_processed['Match_Source'] = ''
        df_processed['Is_Originally_Empty'] = df_processed['Pack form'].isna()
        df_processed['Confidence_Score'] = 0.0
        df_processed['Standardization_Applied'] = False
        
        # 第一步：标准化已存在的剂型
        standardization_count = 0
        for idx, row in df_processed.iterrows():
            if pd.notna(row['Pack form']) and row['Pack form'] != '':
                original_form = row['Pack form']
                standardized_form = self.standardize_pack_form(original_form)
                
                if standardized_form != original_form:
                    df_processed.at[idx, 'Pack form'] = standardized_form
                    df_processed.at[idx, 'Standardization_Applied'] = True
                    standardization_count += 1
        
        # 第二步：处理空的Pack form列
        processed_count = 0
        for idx, row in df_processed.iterrows():
            # 只处理Pack form为空的行
            if pd.isna(row['Pack form']) or row['Pack form'] == '':
                product_text = row['Product']
                detected_forms, matched_texts = self.detect_pack_form(product_text)
                
                if detected_forms:
                    classified_form = self.classify_pack_form(detected_forms)
                    
                    # 实际填充到Pack form列
                    df_processed.at[idx, 'Pack form'] = classified_form
                    
                    # 同时保存到新列
                    df_processed.at[idx, 'Matched_Pack_Form'] = classified_form
                    df_processed.at[idx, 'Match_Source'] = ', '.join(matched_texts)
                    
                    # 计算置信度分数
                    confidence = min(len(detected_forms) / 2.0, 1.0)
                    df_processed.at[idx, 'Confidence_Score'] = confidence
                    
                    processed_count += 1
        
        return df_processed, processed_count, standardization_count
    
    def generate_standardization_report(self, df_processed):
        """
        生成标准化处理报告
        
        Args:
            df_processed (pd.DataFrame): 处理后的DataFrame
            
        Returns:
            dict: 标准化报告
        """
        report = {
            'total_rows': len(df_processed),
            'standardization_applied': df_processed['Standardization_Applied'].sum(),
            'originally_empty': df_processed['Is_Originally_Empty'].sum(),
            'successfully_filled': 0,
            'final_empty': 0,
            'pack_form_distribution': {},
            'standardization_examples': []
        }
        
        # 计算填充统计
        report['successfully_filled'] = report['originally_empty'] - df_processed['Pack form'].isna().sum()
        report['final_empty'] = df_processed['Pack form'].isna().sum()
        
        # 剂型分布
        pack_form_counts = df_processed['Pack form'].value_counts()
        report['pack_form_distribution'] = pack_form_counts.to_dict()
        
        # 标准化示例
        standardized_rows = df_processed[df_processed['Standardization_Applied'] == True]
        if len(standardized_rows) > 0:
            for idx, row in standardized_rows.head(10).iterrows():
                report['standardization_examples'].append({
                    'row': idx + 1,
                    'product': str(row['Product'])[:80] + "..." if len(str(row['Product'])) > 80 else str(row['Product']),
                    'pack_form': row['Pack form']
                })
        
        return report

def pack_form_labeler_app():
    render_app_header("🏷️ 剂型打标工具", "通过匹配产品标题，自动识别剂型并填充到空的Pack form列中")
    col1, col2 = st.columns([2, 1])
    with col1:
        uploaded_file = st.file_uploader(
            "选择您的Excel文件 (.xlsx格式)",
            type=["xlsx"],
            key="pack_form_file"
        )
    with col2:
        save_filename = st.text_input("输出文件名", "labeled_pack_forms.xlsx", key="pack_form_save")
    st.divider()
    if uploaded_file is not None:
        try:
            df_input = _read_excel_cached(uploaded_file)
            st.markdown("#### 文件信息")
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("总行数", len(df_input))
            with col2:
                st.metric("总列数", len(df_input.columns))
            with col3:
                empty_count = df_input['Pack form'].isna().sum() if 'Pack form' in df_input.columns else 0
                st.metric("Pack form空值", empty_count)
            required_columns = ['Pack form', 'Product']
            missing_columns = [col for col in required_columns if col not in df_input.columns]
            if missing_columns:
                st.error(f"文件缺少必要的列: {missing_columns}")
            else:
                st.success("文件格式正确，包含所有必要的列")
                st.markdown("#### 数据预览 (前5行)")
                st.dataframe(df_input.head(), use_container_width=True)
                st.divider()
                execute_btn = st.button("🚀 开始剂型打标", key="pack_form_button", use_container_width=True)
                if execute_btn:
                    with st.spinner("🔄 正在进行剂型智能打标，请稍候..."):
                        try:
                            labeler = PackFormLabeler()
                            df_processed, processed_count, standardization_count = labeler.process_dataframe(df_input)
                            st.success("剂型打标完成！")
                            original_empty_count = (df_input['Pack form'].isna() | (df_input['Pack form'] == '')).sum()
                            final_empty_count = (df_processed['Pack form'].isna() | (df_processed['Pack form'] == '')).sum()
                            successfully_filled_count = original_empty_count - final_empty_count
                            col1, col2, col3, col4, col5 = st.columns(5)
                            with col1:
                                st.metric("原始空值", original_empty_count)
                            with col2:
                                st.metric("成功填充", successfully_filled_count)
                            with col3:
                                st.metric("标准化处理", standardization_count)
                            with col4:
                                st.metric("处理后空值", final_empty_count)
                            with col5:
                                if original_empty_count > 0:
                                    success_rate = successfully_filled_count / original_empty_count * 100
                                    st.metric("成功率", f"{success_rate:.1f}%")
                                else:
                                    st.metric("成功率", "N/A")
                            if standardization_count > 0:
                                st.markdown("#### 标准化处理详情")
                                st.info(f"对 {standardization_count} 行已有剂型进行了标准化处理")
                            st.markdown("#### 剂型分布")
                            pack_form_counts = df_processed['Pack form'].value_counts()
                            st.bar_chart(pack_form_counts)
                            st.markdown("#### 处理结果预览 (前5行)")
                            st.dataframe(df_processed.head(), use_container_width=True)
                            buffer = save_df_to_buffer(df_processed)
                            ts = get_timestamp()
                            out_name = f"labeled_{ts}.xlsx"
                            out_path = os.path.join("/tmp", out_name)
                            save_func = lambda: df_processed.to_excel(out_path, index=False, engine="openpyxl")
                            render_download_section(
                                buffer,
                                out_name,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                "📥 下载打标后的Excel文件",
                                "pack_form",
                                has_save=True,
                                save_func=save_func,
                                save_path=out_path,
                            )
                            st.info("下载的文件包含：原始数据、填充和标准化后的Pack form列，以及新增的匹配信息列")
                        except Exception as e:
                            st.error(f"处理过程中发生错误: {str(e)}")
        except Exception as e:
            st.error(f"读取文件时发生错误: {str(e)}")

# 主应用程序
def main():
    st.set_page_config(page_title=APP_CONFIG["app_title"], layout="wide", page_icon="📊", initial_sidebar_state="collapsed")
    st.markdown("""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
        html, body, [class*="css"] {font-family: 'Inter', 'Segoe UI', sans-serif;}
        .main {background: linear-gradient(135deg, #f5f7fa 0%, #e9ecef 100%);}
        h1, h2, h3, h4, h5, h6 {color: #ffffff !important; font-weight: 600 !important;}
        .stButton > button {
            background: linear-gradient(135deg, #00a6e4 0%, #0088c2 100%);
            color: white; border: none; border-radius: 8px; padding: 0.6rem 1.5rem;
            font-weight: 600; font-size: 15px; transition: all 0.3s ease;
            box-shadow: 0 4px 6px rgba(0, 166, 228, 0.2);
        }
        .stButton > button:hover {
            background: linear-gradient(135deg, #0088c2 0%, #006a99 100%);
            box-shadow: 0 6px 12px rgba(0, 166, 228, 0.3); transform: translateY(-2px);
        }
        .stDownloadButton > button {background: linear-gradient(135deg, #00a6e4 0%, #0088c2 100%);
            color: white; border: none; border-radius: 8px; padding: 0.6rem 1.5rem;
            font-weight: 600; font-size: 15px; transition: all 0.3s ease;
            box-shadow: 0 4px 6px rgba(0, 166, 228, 0.2);
        }
        .stDownloadButton > button:hover {
            background: linear-gradient(135deg, #0088c2 0%, #006a99 100%);
            box-shadow: 0 6px 12px rgba(0, 166, 228, 0.3); transform: translateY(-2px);
        }
        .stFileUploader {background: white; border-radius: 10px; padding: 1.5rem; box-shadow: 0 2px 8px rgba(0,0,0,0.08);}
        [data-testid="stFileUploadDropzone"] {border: 2px dashed #00a6e4; border-radius: 8px; background: #f8fcff;}
        .stTextInput > div > div > input, .stTextArea > div > div > textarea {
            border: 2px solid #e0e0e0; border-radius: 8px; padding: 0.6rem; transition: all 0.3s ease; font-size: 14px;
        }
        .stTextInput > div > div > input:focus, .stTextArea > div > div > textarea:focus {
            border-color: #00a6e4; box-shadow: 0 0 0 3px rgba(0, 166, 228, 0.1);
        }
        .stProgress > div > div > div > div {background: linear-gradient(90deg, #00a6e4 0%, #0088c2 100%);}
        .stSuccess {background: linear-gradient(135deg, #d4edda 0%, #c3e6cb 100%); border-left: 4px solid #28a745; border-radius: 8px; padding: 1rem;}
        .stError {background: linear-gradient(135deg, #f8d7da 0%, #f5c6cb 100%); border-left: 4px solid #dc3545; border-radius: 8px; padding: 1rem;}
        .stWarning {background: linear-gradient(135deg, #fff3cd 0%, #ffeaa7 100%); border-left: 4px solid #ffc107; border-radius: 8px; padding: 1rem;}
        .stInfo {background: linear-gradient(135deg, #d1ecf1 0%, #bee5eb 100%); border-left: 4px solid #00a6e4; border-radius: 8px; padding: 1rem;}
        div[data-testid="column"] {background: white; padding: 1rem; border-radius: 10px; box-shadow: 0 2px 8px rgba(0,0,0,0.05);}
        .js-plotly-plot {border-radius: 10px; box-shadow: 0 4px 12px rgba(0,0,0,0.1);}
    </style>
    """, unsafe_allow_html=True)
    st.markdown(f"""
    <div style="background: linear-gradient(135deg, #00a6e4 0%, #0088c2 100%); padding: 2.5rem 2rem; border-radius: 15px; margin-bottom: 2rem; box-shadow: 0 8px 16px rgba(0,0,0,0.15);">
        <h1 style="color: white; margin: 0; font-size: 2.5rem; font-weight: 700;">📊 市场洞察小程序</h1>
        <div style="display: flex; gap: 2rem; margin-top: 1rem; flex-wrap: wrap;">
            <span style="color: rgba(255,255,255,0.95); font-size: 14px;"><strong>版本:</strong> {APP_CONFIG["version"]}</span>
            <span style="color: rgba(255,255,255,0.95); font-size: 14px;"><strong>作者:</strong> {APP_CONFIG["author"]}</span>
            <span style="color: rgba(255,255,255,0.95); font-size: 14px;"><strong>公司:</strong> {APP_CONFIG["company"]}</span>
            <span style="color: rgba(255,255,255,0.95); font-size: 14px;"><strong>联系:</strong> {APP_CONFIG["contact"]}</span>
        </div>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("""
    <div style="background: white; padding: 1.5rem; border-radius: 10px; margin-bottom: 2rem; box-shadow: 0 2px 8px rgba(0,0,0,0.08);">
        <h3 style="margin-top: 0; color: #333;">🎯 功能导航</h3>
        <p style="color: #666; margin-bottom: 0;">选择下方功能模块开始您的数据分析之旅</p>
    </div>
    """, unsafe_allow_html=True)
    tabs = st.tabs(["📊 合并数据表格", "🔍 搜索流量洞察", "📈 流量可视化分析", "🧹 数据清理工具", "🏷️ 剂型打标工具"])
    with tabs[0]:
        merge_data_app()
    with tabs[1]:
        search_insight_app()
    with tabs[2]:
        search_insight_viz_app()
    with tabs[3]:
        data_clean_app()
    with tabs[4]:
        pack_form_labeler_app()
    st.divider()
    st.markdown("""
    <div style="text-align: center; color: #666; padding: 2rem 0;">
        <p style="margin: 0;">© Anker Oceanwing Inc. | 海翼IDC团队</p>
        <p style="margin: 0.5rem 0 0 0; font-size: 13px;">市场洞察小程序 v1.2.0 - 让数据分析更简单</p>
    </div>
    """, unsafe_allow_html=True)

if __name__ == "__main__":
    main()
