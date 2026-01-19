import streamlit as st
import os
import re
from openpyxl import load_workbook
import tempfile
import zipfile
from io import BytesIO
import time

# é¡µé¢é…ç½®
st.set_page_config(page_title="Excel æ•°æ®è¯æ€§æ‰“æ ‡å·¥å…·", page_icon="ğŸ“Š", layout="wide")

# CSSæ ·å¼
st.markdown("""
<style>
:root {--primary-color: #00a6e4;}
#MainMenu, footer {visibility: hidden;}
.main-title {color: #00a6e4; text-align: center; font-size: 2.5rem; font-weight: bold; margin-bottom: 0.5rem;}
.sub-title {color: #666; text-align: center; font-size: 1rem; margin-bottom: 2rem;}
.stButton > button {background-color: #00a6e4; color: white; border: none; border-radius: 8px; padding: 0.5rem 2rem; font-weight: bold;}
.stButton > button:hover {background-color: #0088bb; box-shadow: 0 4px 8px rgba(0, 166, 228, 0.3);}
.info-box {background: linear-gradient(135deg, #e6f7ff 0%, #f0f9ff 100%); border-left: 4px solid #00a6e4; padding: 1rem; border-radius: 8px; margin: 1rem 0;}
.stat-card {background: white; border-radius: 10px; padding: 1.5rem; box-shadow: 0 2px 8px rgba(0, 166, 228, 0.1); border-top: 3px solid #00a6e4; text-align: center;}
.stat-number {font-size: 2rem; font-weight: bold; color: #00a6e4;}
.stat-label {color: #666; font-size: 0.9rem; margin-top: 0.5rem;}
.log-container {background-color: #f8f9fa; border: 1px solid #e0e0e0; border-radius: 8px; padding: 1rem; max-height: 400px; overflow-y: auto; font-family: monospace; font-size: 0.85rem;}
.log-entry {padding: 0.25rem 0; border-bottom: 1px solid #e8e8e8;}
.stProgress > div > div > div {background-color: #00a6e4;}
</style>
""", unsafe_allow_html=True)

# åˆå§‹åŒ– session state
if 'logs' not in st.session_state:
    st.session_state.logs = []
if 'processed' not in st.session_state:
    st.session_state.processed = False
if 'authenticated' not in st.session_state:
    st.session_state.authenticated = False

def add_log(message):
    """æ·»åŠ æ—¥å¿—"""
    st.session_state.logs.append(f"[{time.strftime('%H:%M:%S')}] {message}")

def check_password():
    """éªŒè¯å¯†ç """
    def password_entered():
        st.session_state.password_attempted = True
        if st.session_state["password"] == "owblueland2026":
            st.session_state.authenticated = True
            del st.session_state["password"]
        else:
            st.session_state.authenticated = False
    
    if not st.session_state.authenticated:
        st.markdown('<h1 class="main-title">ğŸ” ç³»ç»Ÿç™»å½•</h1>', unsafe_allow_html=True)
        st.markdown('<p class="sub-title">è¯·è¾“å…¥è®¿é—®å¯†ç </p>', unsafe_allow_html=True)
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            st.markdown("<br><br>", unsafe_allow_html=True)
            st.text_input("å¯†ç ", type="password", key="password", on_change=password_entered, placeholder="è¯·è¾“å…¥å¯†ç ...")
            if st.session_state.get("password_attempted", False) and not st.session_state.authenticated:
                st.error("âŒ å¯†ç é”™è¯¯ï¼Œè¯·é‡è¯•")
            st.markdown('<div style="text-align: center; margin-top: 20px; color: #666;"><p>ğŸ”’ æ­¤ç³»ç»Ÿä»…ä¾›æˆæƒç”¨æˆ·ä½¿ç”¨</p><p style="color: #00a6e4;">è¯·è”ç³»ç®¡ç†å‘˜è·å–è®¿é—®å¯†ç </p></div>', unsafe_allow_html=True)
        return False
    return True

def process_files(data_files, match_file):
    """å¤„ç†æ–‡ä»¶çš„ä¸»å‡½æ•° (å·²ä¿®æ”¹)"""
    st.session_state.logs = []
    errors = []
    processed_files = []
    
    try:
        # åŠ è½½åŒ¹é…æ–‡ä»¶ (ä¿®æ”¹ç‚¹ï¼šè¯»å–ä¸¤åˆ—)
        add_log("ğŸ”„ å¼€å§‹åŠ è½½åŒ¹é…æ–‡ä»¶...")
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            tmp.write(match_file.read())
            match_file_path = tmp.name
        
        match_wb = load_workbook(match_file_path)
        match_ws = match_wb.active
        
        # æ„å»ºé›†åˆ
        brand_asin_set = set()
        competitor_brand_set = set()
        
        for row in match_ws.iter_rows(min_row=1, max_col=2, values_only=True):
            # ç¬¬ä¸€åˆ—ï¼šå“ç‰Œ ASIN
            if row[0]:
                brand_asin_set.add(str(row[0]).lower().replace(" ", ""))
            # ç¬¬äºŒåˆ—ï¼šç«å“å“ç‰Œ
            if row[1]:
                competitor_brand_set.add(str(row[1]).lower().strip())
        
        match_wb.close()
        os.unlink(match_file_path)
        
        # è¿™é‡Œä¸ºäº†åŒ¹é… Non-brand ç»†åˆ†é€»è¾‘ï¼Œæˆ‘ä»¬éœ€è¦æŠŠç«å“å“ç‰Œè½¬ä¸ºæ­£åˆ™æ¨¡å¼ï¼ˆå¤„ç†ç©ºæ ¼ï¼‰
        # ç›´æ¥å­˜å‚¨å¤„ç†åçš„å­—ç¬¦ä¸²ç”¨äº in åˆ¤æ–­ï¼Œæˆ–è€…å­˜å‚¨ä¸ºæ­£åˆ™æ¨¡å¼
        # ä¸ºäº†ç®€å•é«˜æ•ˆï¼Œæˆ‘ä»¬å­˜å‚¨ä¸ºå°å†™ä¸”æ— ç©ºæ ¼çš„ç‰ˆæœ¬ç”¨äº in åˆ¤æ–­
        processed_competitor_brands = {brand.replace(" ", "") for brand in competitor_brand_set}
        
        add_log(f"âœ… åŒ¹é…æ–‡ä»¶åŠ è½½å®Œæˆ (å…± {len(brand_asin_set)} ä¸ª Blueland ASIN, {len(processed_competitor_brands)} ä¸ªç«å“å“ç‰Œ)")
        
        # åˆ›å»ºè¿›åº¦æ¡
        progress_bar = st.progress(0)
        status_text = st.empty()
        total_files = len(data_files)
        
        # å¤„ç†æ¯ä¸ªæ•°æ®æ–‡ä»¶
        for idx, data_file in enumerate(data_files):
            try:
                status_text.text(f"æ­£åœ¨å¤„ç†: {data_file.name} ({idx+1}/{total_files})")
                add_log(f"ğŸ“„ å¼€å§‹å¤„ç†æ–‡ä»¶: {data_file.name}")
                
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                    tmp.write(data_file.read())
                    data_file_path = tmp.name
                
                wb = load_workbook(data_file_path)
                ws_original = wb.active
                
                # åˆ›å»ºæ–°sheet
                new_sheet_name = "è¯æ€§æ‰“æ ‡"
                if new_sheet_name in wb.sheetnames:
                    wb.remove(wb[new_sheet_name])
                new_ws = wb.create_sheet(title=new_sheet_name)
                
                # æ”¶é›†æ•°æ® (ä¿®æ”¹ç‚¹ï¼šåªè¯»å–ç¬¬1åˆ— Targeting)
                data_rows = []
                for row in ws_original.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
                    col1_val = str(row[0]).lower().replace(" ", "") if row[0] else ""
                    # åŸå§‹å€¼ç”¨äºåˆ¤æ–­æ˜¯å¦åŒ…å«ç«å“å“ç‰Œï¼ˆéœ€è¦ä¿ç•™ç©ºæ ¼ä¿¡æ¯åšåˆ¤æ–­ï¼Œæˆ–è€…ç»Ÿä¸€å¤„ç†ï¼‰
                    # è¿™é‡Œæˆ‘ä»¬ç”¨åŸå§‹å°å†™å€¼åšç«å“å“ç‰ŒåŒ¹é…
                    raw_val = str(row[0]).lower() if row[0] else ""
                    data_rows.append([col1_val, raw_val])
                
                add_log(f"ğŸ“‹ å¤åˆ¶æ•°æ®å®Œæˆ (å…± {len(data_rows)} è¡Œ)")
                
                # å†™å…¥è¡¨å¤´
                new_ws.append(["", "è¯æ€§"]) # ä¿®æ”¹è¡¨å¤´
                
                # è®¡ç®—æ ‡ç­¾å¹¶å†™å…¥
                for clean_val, raw_val in data_rows:
                    # åˆ¤æ–­æ˜¯å¦æ˜¯ ASIN æ ¼å¼ (B0å¼€å¤´)
                    is_b0_pattern = bool(re.match(r'^b0[0-9a-zA-Z]{8}$', clean_val))
                    
                    if not is_b0_pattern:
                        # è¿™æ˜¯å…³é”®è¯é€»è¾‘
                        if "blueland" in clean_val:
                            label = "Brand KW"
                        else:
                            # ç»†åˆ† Non-brand
                            # æ£€æŸ¥ raw_val (åŸå§‹å°å†™) æ˜¯å¦åŒ…å«ä»»ä½•ç«å“å“ç‰Œ
                            # è¿™é‡Œéœ€è¦å¤„ç†ç«å“å“ç‰Œä¸­çš„ç©ºæ ¼ï¼Œæ¯”å¦‚ "target" åº”è¯¥åŒ¹é… "target" æˆ– "target store"
                            matched_competitor = False
                            for comp_brand in competitor_brand_set:
                                # å°†ç«å“å“ç‰Œå’Œæœç´¢è¯ä¸­çš„ç©ºæ ¼éƒ½è€ƒè™‘è¿›å»
                                # ç®€å•åšæ³•ï¼šæ£€æŸ¥ç«å“å“ç‰Œï¼ˆå»é™¤ç©ºæ ¼ï¼‰æ˜¯å¦åœ¨æœç´¢è¯ï¼ˆå»é™¤ç©ºæ ¼ï¼‰ä¸­
                                # æˆ–è€…æ£€æŸ¥ç«å“å“ç‰Œï¼ˆå¸¦ç©ºæ ¼ï¼‰æ˜¯å¦åœ¨æœç´¢è¯ä¸­
                                comp_clean = comp_brand.replace(" ", "")
                                raw_clean = raw_val.replace(" ", "")
                                if comp_clean in raw_clean:
                                    matched_competitor = True
                                    break
                                # æˆ–è€…ä½œä¸ºå®Œæ•´è¯åŒ¹é…ï¼ˆé˜²æ­¢ target åŒ¹é…åˆ° targettï¼‰
                                # è¿™é‡Œé‡‡ç”¨ç®€å•çš„åŒ…å«é€»è¾‘ï¼Œå¦‚æœéœ€è¦æ›´ä¸¥æ ¼ï¼Œå¯ä»¥ç”¨æ­£åˆ™ \b
                            if matched_competitor:
                                label = "CMP KW"
                            else:
                                label = "Cate KW"
                    else:
                        # è¿™æ˜¯ ASIN é€»è¾‘
                        if clean_val in brand_asin_set:
                            label = "Brand PAT"
                        else:
                            label = "CMP PAT"
                    
                    new_ws.append([clean_val, label])
                
                wb.save(data_file_path)
                wb.close()
                
                with open(data_file_path, 'rb') as f:
                    processed_files.append((data_file.name, f.read()))
                
                os.unlink(data_file_path)
                add_log(f"âœ… æ–‡ä»¶ {data_file.name} å¤„ç†å®Œæˆ")
                
            except Exception as e:
                error_msg = f"âŒ å¤„ç†æ–‡ä»¶ {data_file.name} æ—¶å‡ºé”™: {str(e)}"
                errors.append(error_msg)
                add_log(error_msg)
            
            progress_bar.progress((idx + 1) / total_files)
        
        status_text.text("âœ… æ‰€æœ‰æ–‡ä»¶å¤„ç†å®Œæˆï¼")
        return processed_files, errors
        
    except Exception as e:
        add_log(f"âŒ å‘ç”Ÿé”™è¯¯: {str(e)}")
        return [], [str(e)]

# å¯†ç éªŒè¯
if not check_password():
    st.stop()

# ä¸»ç•Œé¢
st.markdown('<h1 class="main-title">ğŸ“Š Excel æ•°æ®è¯æ€§æ‰“æ ‡å·¥å…· (å®šåˆ¶ç‰ˆ)</h1>', unsafe_allow_html=True)
st.markdown('<p class="sub-title">æ‰¹é‡å¤„ç† Excel æ–‡ä»¶ï¼Œè‡ªåŠ¨è¿›è¡Œè¯æ€§æ ‡æ³¨ | v2.1 (ç«å“ç»†åˆ†ç‰ˆ)</p>', unsafe_allow_html=True)

# ä¾§è¾¹æ 
with st.sidebar:
    st.markdown("### ğŸ“– ä½¿ç”¨è¯´æ˜")
    st.markdown('<div class="info-box"><b>æ“ä½œæ­¥éª¤ï¼š</b><br>1ï¸âƒ£ ä¸Šä¼ åŒ…å«æ•°æ®çš„ Excel æ–‡ä»¶ï¼ˆå¯å¤šä¸ªï¼‰<br>2ï¸âƒ£ ä¸Šä¼ åŒ…å«ä¸¤åˆ—çš„åŒ¹é…æ–‡ä»¶<br>3ï¸âƒ£ ç‚¹å‡»"å¼€å§‹å¤„ç†"æŒ‰é’®<br>4ï¸âƒ£ ç­‰å¾…å¤„ç†å®Œæˆå¹¶ä¸‹è½½ç»“æœ</div>', unsafe_allow_html=True)
    
    st.markdown("### ğŸ“‹ æ–‡ä»¶æ ¼å¼è¦æ±‚")
    with st.expander("ğŸ“ æ•°æ®æ–‡ä»¶æ ¼å¼"):
        st.markdown("**æ–‡ä»¶ç±»å‹**: `.xlsx`\n\n**åˆ—ç»“æ„**:\n- **ç¬¬1åˆ— (Targeting)**: å…³é”®è¯æˆ– ASIN\n\n**æ³¨æ„**:\n- ä¸å†éœ€è¦ Campaign Type åˆ—")
    
    with st.expander("ğŸ” åŒ¹é…æ–‡ä»¶æ ¼å¼ (ä¿®æ”¹)"):
        st.markdown("**æ–‡ä»¶ç±»å‹**: `.xlsx`\n\n**åˆ—ç»“æ„**:\n- **ç¬¬1åˆ—**: Blueland å“ç‰Œ ASIN\n- **ç¬¬2åˆ—**: ç«å“å“ç‰Œåç§° (Keywords)\n\n**ç”¨é€”**:\n- ç¬¬1åˆ—ç”¨äºåˆ¤æ–­ Brand PAT\n- ç¬¬2åˆ—ç”¨äºåˆ¤æ–­ CMP KW")
    
    st.markdown("### ğŸ·ï¸ æ ‡æ³¨è§„åˆ™ (ä¿®æ”¹)")
    st.markdown('<div style="font-size: 0.9rem; line-height: 1.8;"><b>å…³é”®è¯ç±»å‹ï¼š</b><br>ğŸ”¹ <b>Brand KW</b>: åŒ…å« "blueland" çš„å…³é”®è¯<br>ğŸ”¹ <b>CMP KW</b>: ä¸åŒ…å« bluelandï¼Œä½†åŒ…å«åŒ¹é…æ–‡ä»¶ä¸­å®šä¹‰çš„ç«å“å“ç‰Œçš„å…³é”®è¯<br>ğŸ”¹ <b>Cate KW</b>: æ—¢ä¸åŒ…å« blueland ä¹Ÿä¸åŒ…å«ç«å“å“ç‰Œçš„æ™®é€šå“ç±»è¯<br><br><b>ASIN ç±»å‹ï¼š</b><br>ğŸ”¹ <b>Brand PAT</b>: åŒ¹é…æ–‡ä»¶ç¬¬1åˆ—ä¸­çš„ ASIN<br>ğŸ”¹ <b>CMP PAT</b>: éå“ç‰Œ ASIN (ç«å“ ASIN)</div>', unsafe_allow_html=True)

# æ–‡ä»¶ä¸Šä¼ 
st.markdown("## ğŸ“¤ æ–‡ä»¶ä¸Šä¼ ")
col1, col2 = st.columns(2)

with col1:
    st.markdown("### ğŸ“ æ•°æ®æ–‡ä»¶")
    st.markdown('<div style="background-color: #f0f9ff; padding: 10px; border-radius: 5px; margin-bottom: 0px;"><small><b>æ ¼å¼è¦æ±‚ï¼š</b><br>â€¢ æ–‡ä»¶æ ¼å¼ï¼š<code>.xlsx</code><br>â€¢ <b>ä»…éœ€ç¬¬1åˆ— (Targeting)</b>ï¼šåŒ…å«å…³é”®è¯æˆ– ASIN<br>â€¢ <b>ç§»é™¤ç¬¬5åˆ—</b>ï¼šä¸å†éœ€è¦å¹¿å‘Šæ´»åŠ¨ç±»å‹</small></div>', unsafe_allow_html=True)
    data_files = st.file_uploader("é€‰æ‹©è¦å¤„ç†çš„ Excel æ–‡ä»¶ï¼ˆå¯å¤šé€‰ï¼‰", type=['xlsx'], accept_multiple_files=True, key="data_files")
    if data_files:
        st.success(f"âœ… å·²é€‰æ‹© {len(data_files)} ä¸ªæ–‡ä»¶")

with col2:
    st.markdown("### ğŸ” åŒ¹é…æ–‡ä»¶")
    st.markdown('<div style="background-color: #fff5e6; padding: 10px; border-radius: 5px; margin-bottom: 0px;"><small><b>æ ¼å¼è¦æ±‚ï¼š</b><br>â€¢ æ–‡ä»¶æ ¼å¼ï¼š<code>.xlsx</code><br>â€¢ <b>ç¬¬1åˆ—</b>ï¼šBlueland å“ç‰Œ ASIN<br>â€¢ <b>ç¬¬2åˆ—</b>ï¼šç«å“å“ç‰Œè¯ (å¦‚ target, amazon ç­‰)</small></div>', unsafe_allow_html=True)
    match_file = st.file_uploader("é€‰æ‹©åŒ…å« ASIN å’Œç«å“çš„åŒ¹é…æ–‡ä»¶", type=['xlsx'], key="match_file")
    if match_file:
        st.success(f"âœ… å·²é€‰æ‹©: {match_file.name}")

st.markdown("---")

# å¤„ç†æŒ‰é’®
col_btn1, col_btn2, col_btn3 = st.columns([1, 1, 1])
with col_btn2:
    if st.button("ğŸš€ å¼€å§‹å¤„ç†", disabled=not (data_files and match_file), use_container_width=True):
        with st.spinner("æ­£åœ¨å¤„ç†ä¸­ï¼Œè¯·ç¨å€™..."):
            processed_files, errors = process_files(data_files, match_file)
            st.session_state.processed = True
            st.session_state.processed_files = processed_files
            st.session_state.errors = errors

# æ˜¾ç¤ºå¤„ç†ç»“æœ
if st.session_state.processed and 'processed_files' in st.session_state:
    st.markdown("---")
    col_stat1, col_stat2, col_stat3 = st.columns(3)
    with col_stat1:
        st.markdown(f'<div class="stat-card"><div class="stat-number">{len(st.session_state.processed_files)}</div><div class="stat-label">æˆåŠŸå¤„ç†</div></div>', unsafe_allow_html=True)
    with col_stat2:
        st.markdown(f'<div class="stat-card"><div class="stat-number">{len(st.session_state.errors)}</div><div class="stat-label">å¤„ç†å¤±è´¥</div></div>', unsafe_allow_html=True)
    with col_stat3:
        st.markdown(f'<div class="stat-card"><div class="stat-number">{len(data_files)}</div><div class="stat-label">æ€»æ–‡ä»¶æ•°</div></div>', unsafe_allow_html=True)
    
    st.markdown("<br>", unsafe_allow_html=True)
    
    # ä¸‹è½½æŒ‰é’®
    if st.session_state.processed_files:
        if len(st.session_state.processed_files) == 1:
            filename, content = st.session_state.processed_files[0]
            st.download_button("â¬‡ï¸ ä¸‹è½½å¤„ç†åçš„æ–‡ä»¶", data=content, file_name=filename, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        else:
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for filename, content in st.session_state.processed_files:
                    zip_file.writestr(filename, content)
            st.download_button("â¬‡ï¸ ä¸‹è½½æ‰€æœ‰å¤„ç†åçš„æ–‡ä»¶ (ZIP)", data=zip_buffer.getvalue(), file_name="processed_files.zip", mime="application/zip", use_container_width=True)
    
    if st.session_state.errors:
        with st.expander("âš ï¸ æŸ¥çœ‹é”™è¯¯è¯¦æƒ…"):
            for error in st.session_state.errors:
                st.error(error)

# æ—¥å¿—æ˜¾ç¤º
if st.session_state.logs:
    st.markdown("---")
    st.markdown("### ğŸ“‹ å¤„ç†æ—¥å¿—")
    log_html = '<div class="log-container">'
    for log in st.session_state.logs:
        log_html += f'<div class="log-entry">{log}</div>'
    log_html += '</div>'
    st.markdown(log_html, unsafe_allow_html=True)

# é¡µè„š
st.markdown("---")
st.markdown('<div style="text-align: center; color: #666; font-size: 0.85rem;"><p>ğŸ’¡ æç¤ºï¼šç¨‹åºä¼šè‡ªåŠ¨è·³è¿‡æŸåçš„æ–‡ä»¶å¹¶ç»§ç»­å¤„ç†å…¶ä»–æ–‡ä»¶</p><p style="color: #00a6e4;">Powered by Streamlit | Â© 2024</p></div>', unsafe_allow_html=True)
