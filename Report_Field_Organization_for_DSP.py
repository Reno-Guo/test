import io
import os
import re
from datetime import date, datetime

import streamlit as st


DATE_HEADERS = ["Year", "Day", "Week", "Month"]
ORDER_FIXED_HEADERS = [
    "Order Time",
    "Country",
    "Brand",
    "Strategy / Product Category",
    "Order Funnel",
]
LINE_FIXED_HEADERS = ["Device", "Inventory", "Audience", "Audience Details"]
CREATIVE_FIXED_HEADERS = [
    "Creative Creation Time",
    "Creative Product Brand",
    "Creative Product Category",
    "Creative type",
    "Creative Resolution",
    "Creative ASIN",
    "Creative Name",
]
DATE_NUMBER_FORMAT = "mmm dd, yyyy"
PRIMARY_COLOR = "#00a6e4"
WEEK_MODE_MONDAY_TO_SUNDAY = "monday_to_sunday"
WEEK_MODE_SUNDAY_TO_SATURDAY = "sunday_to_saturday"


def normalize_header(value):
    return str(value).strip().lower() if value is not None else ""


def rename_header(header):
    if header is None:
        return ""

    text = str(header).strip()
    lowered = text.lower()
    if lowered == "total sales" or lowered.startswith("total sales "):
        return "Total Sales"
    if lowered == "sales" or lowered.startswith("sales "):
        return "Sales"
    return text


def string_value(value):
    if value is None:
        return ""
    return str(value)


def split_and_trim(value, delimiter):
    text = string_value(value)
    if text == "":
        return []
    return [part.strip() for part in text.split(delimiter)]


def find_header(headers, candidates):
    normalized_map = {normalize_header(header): index for index, header in enumerate(headers)}
    for candidate in candidates:
        match_index = normalized_map.get(normalize_header(candidate))
        if match_index is not None:
            return headers[match_index], match_index
    return None, None


def excel_weeknum(date_value, week_mode=WEEK_MODE_SUNDAY_TO_SATURDAY):
    if week_mode == WEEK_MODE_MONDAY_TO_SUNDAY:
        return int(date_value.strftime("%W")) + 1
    return int(date_value.strftime("%U")) + 1


def week_code(date_value, week_mode=WEEK_MODE_SUNDAY_TO_SATURDAY):
    return f"W{date_value.strftime('%y')}{excel_weeknum(date_value, week_mode):02d}"


def parse_date_value(value):
    if value is None or value == "":
        return None

    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        return None

    for fmt in ("%b %d, %Y", "%Y/%m/%d", "%Y-%m-%d", "%m/%d/%Y", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def build_order_headers(total_columns):
    headers = ORDER_FIXED_HEADERS[:]
    custom_count = max(total_columns - len(ORDER_FIXED_HEADERS), 0)
    for index in range(custom_count):
        headers.append(f"Order Custom Field {index + 1}")
    return headers


def build_line_headers(total_columns):
    headers = LINE_FIXED_HEADERS[:]
    custom_count = max(total_columns - len(LINE_FIXED_HEADERS), 0)
    for index in range(custom_count):
        headers.append(f"Line Custom Field {index + 1}")
    return headers


def build_creative_headers(total_columns):
    headers = CREATIVE_FIXED_HEADERS[:]
    custom_count = max(total_columns - len(CREATIVE_FIXED_HEADERS), 0)
    for index in range(custom_count):
        headers.append(f"Creative Custom Field {index + 1}")
    return headers


def parse_order_fields(value, total_columns):
    parts = split_and_trim(value, "_")
    result = [
        parts[0] if len(parts) > 0 else "",
        parts[1] if len(parts) > 1 else "",
        parts[2] if len(parts) > 2 else "",
        parts[3] if len(parts) > 3 else "",
        parts[4] if len(parts) > 4 else "",
    ]

    custom_count = max(total_columns - len(ORDER_FIXED_HEADERS), 0)
    for index in range(custom_count):
        part_index = 5 + index
        result.append(parts[part_index] if len(parts) > part_index else "")

    return result


def parse_line_fields(value, total_columns):
    parts = split_and_trim(value, "|")
    audience_details = parts[2] if len(parts) > 2 else ""
    audience = audience_details.split("-", 1)[0].rstrip() if audience_details else ""

    result = [
        parts[0] if len(parts) > 0 else "",
        parts[1] if len(parts) > 1 else "",
        audience,
        audience_details,
    ]

    custom_count = max(total_columns - len(LINE_FIXED_HEADERS), 0)
    for index in range(custom_count):
        part_index = 3 + index
        result.append(parts[part_index] if len(parts) > part_index else "")

    return result


def parse_creative_type_and_resolution(value):
    text = string_value(value).strip()
    if not text:
        return "", ""

    if re.fullmatch(r"[A-Za-z ]+", text):
        return text, ""

    first_space = text.find(" ")
    if first_space == -1:
        return text, ""

    creative_type = text[:first_space].strip()
    resolution = text[first_space + 1 :].strip()
    return creative_type, resolution


def parse_creative_fields(value, total_columns):
    parts = split_and_trim(value, "|")
    creative_type, resolution = parse_creative_type_and_resolution(parts[3] if len(parts) > 3 else "")

    result = [
        parts[0] if len(parts) > 0 else "",
        parts[1] if len(parts) > 1 else "",
        parts[2] if len(parts) > 2 else "",
        creative_type,
        resolution,
        parts[4] if len(parts) > 4 else "",
        parts[5] if len(parts) > 5 else "",
    ]

    custom_count = max(total_columns - len(CREATIVE_FIXED_HEADERS), 0)
    for index in range(custom_count):
        part_index = 6 + index
        result.append(parts[part_index] if len(parts) > part_index else "")

    return result


def derive_date_fields(value, week_mode=WEEK_MODE_SUNDAY_TO_SATURDAY):
    parsed_date = parse_date_value(value)
    if parsed_date is None:
        return "", value if value is not None else "", "", ""

    return (
        parsed_date.year,
        parsed_date,
        week_code(parsed_date, week_mode),
        parsed_date.month,
    )


def count_delimiter(value, delimiter):
    return string_value(value).count(delimiter)


def load_openpyxl():
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as exc:
        raise ImportError("缺少 openpyxl，请先运行：pip3 install openpyxl") from exc

    return Workbook, load_workbook


def build_processed_workbook(workbook, week_mode=WEEK_MODE_SUNDAY_TO_SATURDAY):
    Workbook, _ = load_openpyxl()
    worksheet = workbook.active

    if worksheet.max_row < 1 or worksheet.max_column < 1:
        raise ValueError("当前工作表为空，无法处理。")

    original_headers = [rename_header(cell.value) for cell in worksheet[1]]

    date_header, date_index = find_header(original_headers, ["Date"])
    order_header, order_index = find_header(original_headers, ["Order", "Campaign name"])
    line_header, line_index = find_header(original_headers, ["Line item", "Ad group name"])
    creative_header, creative_index = find_header(original_headers, ["Creative", "Ad name"])

    missing_headers = []
    if date_header is None:
        missing_headers.append("Date")
    if order_header is None:
        missing_headers.append("Order / Campaign name")
    if line_header is None:
        missing_headers.append("Line item / Ad group name")
    if creative_header is None:
        missing_headers.append("Creative / Ad name")

    if missing_headers:
        raise ValueError("缺少必要列：" + "、".join(missing_headers))

    order_max = 0
    line_max = 0
    creative_max = 0
    for row in worksheet.iter_rows(min_row=2):
        order_max = max(order_max, count_delimiter(row[order_index].value, "_"))
        line_max = max(line_max, count_delimiter(row[line_index].value, "|"))
        creative_max = max(creative_max, count_delimiter(row[creative_index].value, "|"))

    order_total_columns = max(order_max + 1, len(ORDER_FIXED_HEADERS))
    line_total_columns = max(line_max + 2, len(LINE_FIXED_HEADERS))
    creative_total_columns = max(creative_max + 2, len(CREATIVE_FIXED_HEADERS))

    new_workbook = Workbook()
    new_worksheet = new_workbook.active
    new_worksheet.title = f"{worksheet.title}_processed"

    order_headers = build_order_headers(order_total_columns)
    line_headers = build_line_headers(line_total_columns)
    creative_headers = build_creative_headers(creative_total_columns)

    prefix_headers = DATE_HEADERS + order_headers + line_headers + creative_headers
    all_headers = prefix_headers + original_headers

    for column_index, header in enumerate(all_headers, start=1):
        new_worksheet.cell(row=1, column=column_index, value=header)

    original_start_column = len(prefix_headers) + 1

    for new_row_index, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        date_value = row[date_index].value
        order_value = row[order_index].value
        line_value = row[line_index].value
        creative_value = row[creative_index].value

        derived_values = []
        derived_values.extend(derive_date_fields(date_value, week_mode))
        derived_values.extend(parse_order_fields(order_value, order_total_columns))
        derived_values.extend(parse_line_fields(line_value, line_total_columns))
        derived_values.extend(parse_creative_fields(creative_value, creative_total_columns))

        for column_index, value in enumerate(derived_values, start=1):
            cell = new_worksheet.cell(row=new_row_index, column=column_index, value=value)
            if column_index == 2 and value not in ("", None):
                cell.number_format = DATE_NUMBER_FORMAT

        for offset, source_cell in enumerate(row, start=original_start_column):
            target_cell = new_worksheet.cell(row=new_row_index, column=offset, value=source_cell.value)
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format

    new_worksheet.freeze_panes = "A2"
    new_worksheet.auto_filter.ref = new_worksheet.dimensions
    return new_workbook, max(worksheet.max_row - 1, 0)


def process_workbook(input_path, output_path, week_mode=WEEK_MODE_SUNDAY_TO_SATURDAY):
    _, load_workbook = load_openpyxl()
    workbook = load_workbook(input_path)
    new_workbook, _ = build_processed_workbook(workbook, week_mode)
    new_workbook.save(output_path)


def process_workbook_bytes(input_bytes, week_mode=WEEK_MODE_SUNDAY_TO_SATURDAY):
    _, load_workbook = load_openpyxl()
    workbook = load_workbook(io.BytesIO(input_bytes))
    new_workbook, processed_rows = build_processed_workbook(workbook, week_mode)
    output_buffer = io.BytesIO()
    new_workbook.save(output_buffer)
    output_buffer.seek(0)
    return output_buffer.getvalue(), processed_rows


def inject_styles():
    st.markdown(
        f"""
        <style>
            :root {{
                --brand: {PRIMARY_COLOR};
            }}
            .stApp {{
                background: linear-gradient(180deg, #f7fcff 0%, #ffffff 45%);
            }}
            h1, h2, h3 {{
                color: var(--brand) !important;
            }}
            .stButton > button,
            .stDownloadButton > button {{
                background-color: var(--brand) !important;
                color: #ffffff !important;
                border: 1px solid var(--brand) !important;
                border-radius: 10px !important;
                font-weight: 600 !important;
            }}
            .stButton > button:hover,
            .stDownloadButton > button:hover {{
                background-color: #008dc2 !important;
                border-color: #008dc2 !important;
            }}
            .block-container {{
                padding-top: 2rem !important;
            }}
        </style>
        """,
        unsafe_allow_html=True,
    )


def streamlit_app():
    st.set_page_config(page_title="广告报表字段拆分", page_icon="📊", layout="centered")
    inject_styles()

    st.title("广告报表字段拆分")
    st.caption("上传 `.xlsx` 文件，自动拆分 Date / Order / Line item / Creative 字段并下载处理结果。")
    week_mode_label = st.selectbox(
        "周定义",
        options=["周一到周日", "周日到周六"],
        index=0,
        help="选择后会影响输出结果中的 Week 周数。",
    )
    week_mode = (
        WEEK_MODE_MONDAY_TO_SUNDAY
        if week_mode_label == "周一到周日"
        else WEEK_MODE_SUNDAY_TO_SATURDAY
    )

    uploaded_file = st.file_uploader("上传广告报表", type=["xlsx"])
    if uploaded_file is None:
        st.info("请先上传一个 Excel 文件。")
        return

    source_name = uploaded_file.name
    base_name = os.path.splitext(source_name)[0]
    output_name = f"{base_name}_processed.xlsx"

    st.write(f"当前文件：`{source_name}`")

    if st.button("开始处理", type="primary", use_container_width=True):
        try:
            with st.spinner("正在处理，请稍候..."):
                output_bytes, processed_rows = process_workbook_bytes(uploaded_file.getvalue(), week_mode)
        except Exception as exc:
            st.error(f"处理失败：{exc}")
            return

        st.success(f"处理完成，共处理 {processed_rows} 行数据。")
        st.download_button(
            label="下载处理后的文件",
            data=output_bytes,
            file_name=output_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )


if __name__ == "__main__":
    streamlit_app()
