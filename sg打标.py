import streamlit as st
import os
import re
from openpyxl import load_workbook
import tempfile
import zipfile
from io import BytesIO
import time

# 页面配置
st.set_page_config(page_title="Excel 数据词性打标工具", page_icon="📊", layout="wide")

# CSS样式
st.markdown("""
<style>
:root {--primary-color: #00a6e4;}
#MainMenu, footer {visibility: hidden;}
.main-title {color: #00a6e4; text-align: center; font-size: 2.5rem; font-weight: bold; margin-bottom: 0.5rem;}
.sub-title {color: #666; text-align: center; font-size: 1rem; margin-bottom: 2rem;}
.stButton > button {background-color: #00a6e4; color: white; border: none; border-radius: 8px; padding: 0.5rem 2rem; font-weight: bold;}
.stButton > button:hover {background-color: #0088bb; box-shadow: 0 4px 8px rgba(0, 166, 228, 0.3);}
.info-box {background: linear-gradient(135deg, #e6f7ff 0%, #f0f9ff 100%); border-left: 4px solid #00a6e4; padding: 1rem; border-radius: 8px; margin: 1rem 0;}
.stat-card {background: white; border-radius: 10px; padding: 1.5rem; box-shadow: 0 2px 8px rgba(0, 166, 228, 0.1); border-top: 3px solid #00a6e4; text-align: center;}
.stat-number {font-size: 2rem; font-weight: bold; color: #00a6e4;}
.stat-label {color: #666; font-size: 0.9rem; margin-top: 0.5rem;}
.log-container {background-color: #f8f9fa; border: 1px solid #e0e0e0; border-radius: 8px; padding: 1rem; max-height: 400px; overflow-y: auto; font-family: monospace; font-size: 0.85rem;}
.log-entry {padding: 0.25rem 0; border-bottom: 1px solid #e8e8e8;}
.stProgress > div > div > div {background-color: #00a6e4;}
</style>
""", unsafe_allow_html=True)

# 初始化 session state
if 'logs' not in st.session_state:
    st.session_state.logs = []
if 'processed' not in st.session_state:
    st.session_state.processed = False
if 'authenticated' not in st.session_state:
    st.session_state.authenticated = False

def add_log(message):
    """添加日志"""
    st.session_state.logs.append(f"[{time.strftime('%H:%M:%S')}] {message}")

def check_password():
    """验证密码"""
    def password_entered():
        st.session_state.password_attempted = True
        if st.session_state["password"] == "owsupergut2026":
            st.session_state.authenticated = True
            del st.session_state["password"]
        else:
            st.session_state.authenticated = False
    
    if not st.session_state.authenticated:
        st.markdown('<h1 class="main-title">🔐 系统登录</h1>', unsafe_allow_html=True)
        st.markdown('<p class="sub-title">请输入访问密码</p>', unsafe_allow_html=True)
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            st.markdown("<br><br>", unsafe_allow_html=True)
            st.text_input("密码", type="password", key="password", on_change=password_entered, placeholder="请输入密码...")
            if st.session_state.get("password_attempted", False) and not st.session_state.authenticated:
                st.error("❌ 密码错误，请重试")
            st.markdown('<div style="text-align: center; margin-top: 20px; color: #666;"><p>🔒 此系统仅供授权用户使用</p><p style="color: #00a6e4;">请联系管理员获取访问密码</p></div>', unsafe_allow_html=True)
        return False
    return True

def process_files(data_files, match_file):
    """处理文件的主函数 (已修改)"""
    st.session_state.logs = []
    errors = []
    processed_files = []
    
    try:
        # 加载匹配文件 (修改点：读取两列)
        add_log("🔄 开始加载匹配文件...")
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            tmp.write(match_file.read())
            match_file_path = tmp.name
        
        match_wb = load_workbook(match_file_path)
        match_ws = match_wb.active
        
        # 构建集合
        brand_asin_set = set()
        competitor_brand_set = set()
        
        for row in match_ws.iter_rows(min_row=1, max_col=2, values_only=True):
            # 第一列：品牌 ASIN
            if row[0]:
                brand_asin_set.add(str(row[0]).lower().replace(" ", ""))
            # 第二列：竞品品牌
            if row[1]:
                competitor_brand_set.add(str(row[1]).lower().strip())
        
        match_wb.close()
        os.unlink(match_file_path)
        
        # 这里为了匹配 Non-brand 细分逻辑，我们需要把竞品品牌转为正则模式（处理空格）
        # 直接存储处理后的字符串用于 in 判断，或者存储为正则模式
        # 为了简单高效，我们存储为小写且无空格的版本用于 in 判断
        processed_competitor_brands = {brand.replace(" ", "") for brand in competitor_brand_set}
        
        add_log(f"✅ 匹配文件加载完成 (共 {len(brand_asin_set)} 个 supergut ASIN, {len(processed_competitor_brands)} 个竞品品牌)")
        
        # 创建进度条
        progress_bar = st.progress(0)
        status_text = st.empty()
        total_files = len(data_files)
        
        # 处理每个数据文件
        for idx, data_file in enumerate(data_files):
            try:
                status_text.text(f"正在处理: {data_file.name} ({idx+1}/{total_files})")
                add_log(f"📄 开始处理文件: {data_file.name}")
                
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                    tmp.write(data_file.read())
                    data_file_path = tmp.name
                
                wb = load_workbook(data_file_path)
                ws_original = wb.active
                
                # 创建新sheet
                new_sheet_name = "词性打标"
                if new_sheet_name in wb.sheetnames:
                    wb.remove(wb[new_sheet_name])
                new_ws = wb.create_sheet(title=new_sheet_name)
                
                # 收集数据 (修改点：只读取第1列 Targeting)
                data_rows = []
                for row in ws_original.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
                    col1_val = str(row[0]).lower().replace(" ", "") if row[0] else ""
                    # 原始值用于判断是否包含竞品品牌（需要保留空格信息做判断，或者统一处理）
                    # 这里我们用原始小写值做竞品品牌匹配
                    raw_val = str(row[0]).lower() if row[0] else ""
                    data_rows.append([col1_val, raw_val])
                
                add_log(f"📋 复制数据完成 (共 {len(data_rows)} 行)")
                
                # 写入表头
                new_ws.append(["", "词性"]) # 修改表头
                
                # 计算标签并写入
                for clean_val, raw_val in data_rows:
                    # 判断是否是 ASIN 格式 (B0开头)
                    is_b0_pattern = bool(re.match(r'^b0[0-9a-zA-Z]{8}$', clean_val))
                    
                    if not is_b0_pattern:
                        # 这是关键词逻辑
                        if "supergut" in clean_val:
                            label = "Brand KW"
                        else:
                            # 细分 Non-brand
                            # 检查 raw_val (原始小写) 是否包含任何竞品品牌
                            # 这里需要处理竞品品牌中的空格，比如 "target" 应该匹配 "target" 或 "target store"
                            matched_competitor = False
                            for comp_brand in competitor_brand_set:
                                # 将竞品品牌和搜索词中的空格都考虑进去
                                # 简单做法：检查竞品品牌（去除空格）是否在搜索词（去除空格）中
                                # 或者检查竞品品牌（带空格）是否在搜索词中
                                comp_clean = comp_brand.replace(" ", "")
                                raw_clean = raw_val.replace(" ", "")
                                if comp_clean in raw_clean:
                                    matched_competitor = True
                                    break
                                # 或者作为完整词匹配（防止 target 匹配到 targett）
                                # 这里采用简单的包含逻辑，如果需要更严格，可以用正则 \b
                            if matched_competitor:
                                label = "CMP KW"
                            else:
                                label = "Cate KW"
                    else:
                        # 这是 ASIN 逻辑
                        if clean_val in brand_asin_set:
                            label = "Brand PAT"
                        else:
                            label = "CMP PAT"
                    
                    new_ws.append([clean_val, label])
                
                wb.save(data_file_path)
                wb.close()
                
                with open(data_file_path, 'rb') as f:
                    processed_files.append((data_file.name, f.read()))
                
                os.unlink(data_file_path)
                add_log(f"✅ 文件 {data_file.name} 处理完成")
                
            except Exception as e:
                error_msg = f"❌ 处理文件 {data_file.name} 时出错: {str(e)}"
                errors.append(error_msg)
                add_log(error_msg)
            
            progress_bar.progress((idx + 1) / total_files)
        
        status_text.text("✅ 所有文件处理完成！")
        return processed_files, errors
        
    except Exception as e:
        add_log(f"❌ 发生错误: {str(e)}")
        return [], [str(e)]

# 密码验证
if not check_password():
    st.stop()

# 主界面
st.markdown('<h1 class="main-title">📊 Excel 数据词性打标工具 (定制版)</h1>', unsafe_allow_html=True)
st.markdown('<p class="sub-title">批量处理 Excel 文件，自动进行词性标注 | v2.1 (竞品细分版)</p>', unsafe_allow_html=True)

# 侧边栏
with st.sidebar:
    st.markdown("### 📖 使用说明")
    st.markdown('<div class="info-box"><b>操作步骤：</b><br>1️⃣ 上传包含数据的 Excel 文件（可多个）<br>2️⃣ 上传包含两列的匹配文件<br>3️⃣ 点击"开始处理"按钮<br>4️⃣ 等待处理完成并下载结果</div>', unsafe_allow_html=True)
    
    st.markdown("### 📋 文件格式要求")
    with st.expander("📁 数据文件格式"):
        st.markdown("**文件类型**: `.xlsx`\n\n**列结构**:\n- **第1列 (Targeting)**: 关键词或 ASIN\n\n**注意**:\n- 不再需要 Campaign Type 列")
    
    with st.expander("🔍 匹配文件格式 (修改)"):
        st.markdown("**文件类型**: `.xlsx`\n\n**列结构**:\n- **第1列**: supergut 品牌 ASIN\n- **第2列**: 竞品品牌名称 (Keywords)\n\n**用途**:\n- 第1列用于判断 Brand PAT\n- 第2列用于判断 CMP KW")
    
    st.markdown("### 🏷️ 标注规则 (修改)")
    st.markdown('<div style="font-size: 0.9rem; line-height: 1.8;"><b>关键词类型：</b><br>🔹 <b>Brand KW</b>: 包含 "supergut" 的关键词<br>🔹 <b>CMP KW</b>: 不包含 supergut，但包含匹配文件中定义的竞品品牌的关键词<br>🔹 <b>Cate KW</b>: 既不包含 supergut 也不包含竞品品牌的普通品类词<br><br><b>ASIN 类型：</b><br>🔹 <b>Brand PAT</b>: 匹配文件第1列中的 ASIN<br>🔹 <b>CMP PAT</b>: 非品牌 ASIN (竞品 ASIN)</div>', unsafe_allow_html=True)

# 文件上传
st.markdown("## 📤 文件上传")
col1, col2 = st.columns(2)

with col1:
    st.markdown("### 📁 数据文件")
    st.markdown('<div style="background-color: #f0f9ff; padding: 10px; border-radius: 5px; margin-bottom: 0px;"><small><b>格式要求：</b><br>• 文件格式：<code>.xlsx</code><br>• <b>仅需第1列 (Targeting)</b>：包含关键词或 ASIN<br>• <b>移除第5列</b>：不再需要广告活动类型</small></div>', unsafe_allow_html=True)
    data_files = st.file_uploader("选择要处理的 Excel 文件（可多选）", type=['xlsx'], accept_multiple_files=True, key="data_files")
    if data_files:
        st.success(f"✅ 已选择 {len(data_files)} 个文件")

with col2:
    st.markdown("### 🔍 匹配文件")
    st.markdown('<div style="background-color: #fff5e6; padding: 10px; border-radius: 5px; margin-bottom: 0px;"><small><b>格式要求：</b><br>• 文件格式：<code>.xlsx</code><br>• <b>第1列</b>：supergut 品牌 ASIN<br>• <b>第2列</b>：竞品品牌词 (如 target, amazon 等)</small></div>', unsafe_allow_html=True)
    match_file = st.file_uploader("选择包含 ASIN 和竞品的匹配文件", type=['xlsx'], key="match_file")
    if match_file:
        st.success(f"✅ 已选择: {match_file.name}")

st.markdown("---")

# 处理按钮
col_btn1, col_btn2, col_btn3 = st.columns([1, 1, 1])
with col_btn2:
    if st.button("🚀 开始处理", disabled=not (data_files and match_file), use_container_width=True):
        with st.spinner("正在处理中，请稍候..."):
            processed_files, errors = process_files(data_files, match_file)
            st.session_state.processed = True
            st.session_state.processed_files = processed_files
            st.session_state.errors = errors

# 显示处理结果
if st.session_state.processed and 'processed_files' in st.session_state:
    st.markdown("---")
    col_stat1, col_stat2, col_stat3 = st.columns(3)
    with col_stat1:
        st.markdown(f'<div class="stat-card"><div class="stat-number">{len(st.session_state.processed_files)}</div><div class="stat-label">成功处理</div></div>', unsafe_allow_html=True)
    with col_stat2:
        st.markdown(f'<div class="stat-card"><div class="stat-number">{len(st.session_state.errors)}</div><div class="stat-label">处理失败</div></div>', unsafe_allow_html=True)
    with col_stat3:
        st.markdown(f'<div class="stat-card"><div class="stat-number">{len(data_files)}</div><div class="stat-label">总文件数</div></div>', unsafe_allow_html=True)
    
    st.markdown("<br>", unsafe_allow_html=True)
    
    # 下载按钮
    if st.session_state.processed_files:
        if len(st.session_state.processed_files) == 1:
            filename, content = st.session_state.processed_files[0]
            st.download_button("⬇️ 下载处理后的文件", data=content, file_name=filename, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        else:
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for filename, content in st.session_state.processed_files:
                    zip_file.writestr(filename, content)
            st.download_button("⬇️ 下载所有处理后的文件 (ZIP)", data=zip_buffer.getvalue(), file_name="processed_files.zip", mime="application/zip", use_container_width=True)
    
    if st.session_state.errors:
        with st.expander("⚠️ 查看错误详情"):
            for error in st.session_state.errors:
                st.error(error)

# 日志显示
if st.session_state.logs:
    st.markdown("---")
    st.markdown("### 📋 处理日志")
    log_html = '<div class="log-container">'
    for log in st.session_state.logs:
        log_html += f'<div class="log-entry">{log}</div>'
    log_html += '</div>'
    st.markdown(log_html, unsafe_allow_html=True)

# 页脚
st.markdown("---")
st.markdown('<div style="text-align: center; color: #666; font-size: 0.85rem;"><p>💡 提示：程序会自动跳过损坏的文件并继续处理其他文件</p><p style="color: #00a6e4;">Powered by Streamlit | © 2024</p></div>', unsafe_allow_html=True)
