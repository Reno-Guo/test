import streamlit as st
import os
import re
from openpyxl import load_workbook
import tempfile
import zipfile
from io import BytesIO
import time

# 页面配置
st.set_page_config(
    page_title="Excel 数据词性打标工具",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# 自定义CSS样式
st.markdown("""
    <style>
    /* 主色调 */
    :root {
        --primary-color: #00a6e4;
    }
    
    /* 隐藏默认的菜单和页脚 */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    
    /* 标题样式 */
    .main-title {
        color: #00a6e4;
        text-align: center;
        font-size: 2.5rem;
        font-weight: bold;
        margin-bottom: 0.5rem;
        text-shadow: 2px 2px 4px rgba(0, 166, 228, 0.1);
    }
    
    .sub-title {
        color: #666;
        text-align: center;
        font-size: 1rem;
        margin-bottom: 2rem;
    }
    
    /* 文件上传区域 */
    .uploadedFile {
        border: 2px dashed #00a6e4 !important;
        border-radius: 10px;
        padding: 1rem;
    }
    
    /* 按钮样式 */
    .stButton > button {
        background-color: #00a6e4;
        color: white;
        border: none;
        border-radius: 8px;
        padding: 0.5rem 2rem;
        font-weight: bold;
        transition: all 0.3s;
    }
    
    .stButton > button:hover {
        background-color: #0088bb;
        box-shadow: 0 4px 8px rgba(0, 166, 228, 0.3);
    }
    
    /* 信息框样式 */
    .info-box {
        background: linear-gradient(135deg, #e6f7ff 0%, #f0f9ff 100%);
        border-left: 4px solid #00a6e4;
        padding: 1rem;
        border-radius: 8px;
        margin: 1rem 0;
    }
    
    /* 统计卡片 */
    .stat-card {
        background: white;
        border-radius: 10px;
        padding: 1.5rem;
        box-shadow: 0 2px 8px rgba(0, 166, 228, 0.1);
        border-top: 3px solid #00a6e4;
        text-align: center;
    }
    
    .stat-number {
        font-size: 2rem;
        font-weight: bold;
        color: #00a6e4;
    }
    
    .stat-label {
        color: #666;
        font-size: 0.9rem;
        margin-top: 0.5rem;
    }
    
    /* 日志区域 */
    .log-container {
        background-color: #f8f9fa;
        border: 1px solid #e0e0e0;
        border-radius: 8px;
        padding: 1rem;
        max-height: 400px;
        overflow-y: auto;
        font-family: monospace;
        font-size: 0.85rem;
    }
    
    .log-entry {
        padding: 0.25rem 0;
        border-bottom: 1px solid #e8e8e8;
    }
    
    /* 进度条 */
    .stProgress > div > div > div {
        background-color: #00a6e4;
    }
    
    /* 侧边栏 */
    .css-1d391kg {
        background-color: #f0f9ff;
    }
    
    /* 成功/错误消息 */
    .success-msg {
        background-color: #d4edda;
        border-left: 4px solid #28a745;
        padding: 1rem;
        border-radius: 8px;
        margin: 1rem 0;
    }
    
    .error-msg {
        background-color: #f8d7da;
        border-left: 4px solid #dc3545;
        padding: 1rem;
        border-radius: 8px;
        margin: 1rem 0;
    }
    </style>
""", unsafe_allow_html=True)

# 密码验证 - 必须在所有内容之前
if not check_password():
    st.stop()  # 如果未通过验证，停止执行后续代码

# 标题
st.markdown('<h1 class="main-title">📊 Excel 数据词性打标工具</h1>', unsafe_allow_html=True)
st.markdown('<p class="sub-title">批量处理 Excel 文件，自动进行词性标注 | v2.0 Streamlit Edition</p>', unsafe_allow_html=True)

# 侧边栏说明
with st.sidebar:
    st.markdown("### 📖 使用说明")
    st.markdown("""
    <div class="info-box">
    <b>操作步骤：</b><br>
    1️⃣ 上传包含数据的 Excel 文件（可多个）<br>
    2️⃣ 上传匹配 ASIN 的 Excel 文件<br>
    3️⃣ 点击"开始处理"按钮<br>
    4️⃣ 等待处理完成并下载结果
    </div>
    """, unsafe_allow_html=True)
    
    st.markdown("### 📋 文件格式要求")
    
    with st.expander("📁 数据文件格式", expanded=False):
        st.markdown("""
        **文件类型**: `.xlsx` (Excel 文件)
        
        **列结构要求**:
        - **第1列**: Query（查询词/ASIN）
        - **第2-4列**: 任意数据
        - **第5列**: Campaign Type（广告类型）
        
        **数据要求**:
        - 第一行为表头
        - 数据从第二行开始
        - Query 可以是关键词或 B0 开头的10位 ASIN
        
        **示例**:
        ```
        | Query      | ... | Campaign Type |
        |------------|-----|---------------|
        | oneplus 12 | ... | Manual        |
        | B09XYZ1234 | ... | Auto          |
        ```
        """)
    
    with st.expander("🔍 匹配文件格式", expanded=False):
        st.markdown("""
        **文件类型**: `.xlsx` (Excel 文件)
        
        **列结构要求**:
        - **第1列**: ASIN 列表（程序只读取第1列）
        - 其他列会被忽略
        
        **数据要求**:
        - ASIN 格式: B0 开头的10位字符
        - 程序会自动去除空格并转为小写
        - 用于判断 Brand PAT 和 CMP PAT
        
        **示例**:
        ```
        | ASIN       |
        |------------|
        | B09ABC1234 |
        | B09DEF5678 |
        ```
        """)
    
    st.info("💡 提示：两种文件都只需要 .xlsx 格式，不支持 .xls 旧版本")
    
    st.markdown("### 🏷️ 标注规则")
    st.markdown("""
    <div style="font-size: 0.9rem; line-height: 1.8;">
    <b>关键词类型：</b><br>
    🔹 <b>Brand KW</b>: 品牌关键词，oneplus相关短语关键词<br>
    🔹 <b>Non-brand KW</b>: 除了oneplus外所有关键词<br>
    <br>
    <b>ASIN 类型：</b><br>
    🔹 <b>Brand PAT</b>: OnePlus相关asin<br>
    🔹 <b>CMP PAT</b>: 竞手Asin（除oneplus相关asin外的所有asin）<br>
    <br>
    <b>自动广告类型：</b><br>
    🔹 <b>Auto KW</b>: OnePlus相关asin，但是有标记自动广告的<br>
    🔹 <b>Auto PAT</b>: 竞手Asin，但是有标记自动广告的<br>
    </div>
    """, unsafe_allow_html=True)
    
    st.markdown("---")
    
    st.markdown("### ⚙️ 处理逻辑")
    st.markdown("""
    <div style="font-size: 0.85rem; background-color: #f8f9fa; padding: 10px; border-radius: 5px;">
    1. 读取数据文件的第1列（Query）和第5列（Campaign Type）<br>
    2. 判断 Query 是否为 B0 开头的10位 ASIN<br>
    3. 对关键词：检查是否包含 "oneplus"<br>
    4. 对ASIN：检查是否在匹配文件中（OnePlus产品）<br>
    5. 对自动广告：检查 Campaign Type 是否包含 "auto"<br>
    6. 在原文件中创建新的 "词性打标" sheet<br>
    7. 保留原始数据，添加标注结果
    </div>
    """, unsafe_allow_html=True)
    
    st.markdown("---")
    st.markdown("**版本**: v2.0")
    st.markdown("**技术栈**: Streamlit + OpenPyxl")

# 初始化 session state
if 'logs' not in st.session_state:
    st.session_state.logs = []
if 'processed' not in st.session_state:
    st.session_state.processed = False
if 'authenticated' not in st.session_state:
    st.session_state.authenticated = False

def add_log(message):
    """添加日志"""
    timestamp = time.strftime("%H:%M:%S")
    st.session_state.logs.append(f"[{timestamp}] {message}")

def check_password():
    """验证密码"""
    def password_entered():
        if st.session_state["password"] == "owoneplus2025":
            st.session_state.authenticated = True
            del st.session_state["password"]  # 删除密码，不保存
        else:
            st.session_state.authenticated = False
    
    if not st.session_state.authenticated:
        st.markdown('<h1 class="main-title">🔐 系统登录</h1>', unsafe_allow_html=True)
        st.markdown('<p class="sub-title">请输入访问密码</p>', unsafe_allow_html=True)
        
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            st.markdown("<br><br>", unsafe_allow_html=True)
            st.text_input(
                "密码",
                type="password",
                key="password",
                on_change=password_entered,
                placeholder="请输入密码..."
            )
            
            if "password" in st.session_state and not st.session_state.authenticated:
                st.error("❌ 密码错误，请重试")
            
            st.markdown("""
                <div style="text-align: center; margin-top: 20px; color: #666; font-size: 0.9rem;">
                    <p>🔒 此系统仅供授权用户使用</p>
                    <p style="color: #00a6e4;">请联系管理员获取访问密码</p>
                </div>
            """, unsafe_allow_html=True)
        
        return False
    
    return True

def process_files(data_files, match_file):
    """处理文件的主函数"""
    st.session_state.logs = []
    errors = []
    processed_files = []
    
    try:
        # 加载匹配文件
        add_log("🔄 开始加载匹配文件...")
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            tmp.write(match_file.read())
            match_file_path = tmp.name
        
        match_wb = load_workbook(match_file_path)
        match_ws = match_wb.active
        match_set = set()
        
        for row in match_ws.iter_rows(min_row=1, min_col=1, max_col=1, values_only=True):
            if row[0]:
                cleaned = str(row[0]).lower().replace(" ", "")
                match_set.add(cleaned)
        
        match_wb.close()
        os.unlink(match_file_path)
        add_log(f"✅ 匹配文件加载完成 (共 {len(match_set)} 个 OnePlus ASIN)")
        
        # 创建进度条
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        total_files = len(data_files)
        
        # 处理每个数据文件
        for idx, data_file in enumerate(data_files):
            try:
                status_text.text(f"正在处理: {data_file.name} ({idx+1}/{total_files})")
                add_log(f"📄 开始处理文件: {data_file.name}")
                
                # 保存上传的文件到临时文件
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                    tmp.write(data_file.read())
                    data_file_path = tmp.name
                
                # 加载工作簿
                wb = load_workbook(data_file_path)
                ws_original = wb.active
                
                # 创建新sheet
                new_sheet_name = "词性打标"
                if new_sheet_name in wb.sheetnames:
                    wb.remove(wb[new_sheet_name])
                new_ws = wb.create_sheet(title=new_sheet_name)
                
                # 收集数据
                data_rows = []
                for row in ws_original.iter_rows(min_row=2, max_col=5, values_only=True):
                    col1_val = str(row[0]).lower().replace(" ", "") if row[0] else ""
                    col5_val = str(row[4]).lower() if len(row) > 4 and row[4] else ""
                    data_rows.append([col1_val, col5_val])
                
                add_log(f"📋 复制数据完成 (共 {len(data_rows)} 行)")
                
                # 写入表头
                new_ws.append(["", "", "词性"])
                
                # 计算标签
                for col1, col2 in data_rows:
                    is_b0_pattern = bool(re.match(r'^b0[0-9a-zA-Z]{8}$', col1))

# 主界面
st.markdown("## 📤 文件上传")

col1, col2 = st.columns(2)

with col1:
    st.markdown("### 📁 数据文件")
    st.markdown("""
    <div style="background-color: #f0f9ff; padding: 10px; border-radius: 5px; margin-bottom: 10px;">
    <small><b>格式要求：</b></small><br>
    <small>• 文件格式：<code>.xlsx</code></small><br>
    <small>• 第1列：Query（关键词/ASIN）</small><br>
    <small>• 第5列：Campaign Type</small><br>
    <small>• 第一行为表头，数据从第二行开始</small>
    </div>
    """, unsafe_allow_html=True)
    
    data_files = st.file_uploader(
        "选择要处理的 Excel 文件（可多选）",
        type=['xlsx'],
        accept_multiple_files=True,
        key="data_files",
        help="支持同时上传多个文件进行批量处理"
    )
    
    if data_files:
        st.success(f"✅ 已选择 {len(data_files)} 个文件")
        with st.expander("📋 查看文件列表"):
            for idx, f in enumerate(data_files, 1):
                file_size = len(f.getvalue()) / 1024  # KB
                st.write(f"{idx}. 📄 {f.name} ({file_size:.1f} KB)")

with col2:
    st.markdown("### 🔍 匹配文件")
    st.markdown("""
    <div style="background-color: #fff5e6; padding: 10px; border-radius: 5px; margin-bottom: 10px;">
    <small><b>格式要求：</b></small><br>
    <small>• 文件格式：<code>.xlsx</code></small><br>
    <small>• 第1列：ASIN 列表</small><br>
    <small>• ASIN 格式：B0 开头的10位字符</small><br>
    <small>• 用于判断 Brand PAT 和 CMP PAT</small>
    </div>
    """, unsafe_allow_html=True)
    
    match_file = st.file_uploader(
        "选择包含 ASIN 的匹配文件（单个）",
        type=['xlsx'],
        key="match_file",
        help="此文件用于匹配 ASIN，判断是否为品牌产品"
    )
    
    if match_file:
        file_size = len(match_file.getvalue()) / 1024  # KB
        st.success(f"✅ 已选择: {match_file.name} ({file_size:.1f} KB)")

st.markdown("---")

# 处理按钮
col_btn1, col_btn2, col_btn3 = st.columns([1, 1, 1])
with col_btn2:
    if st.button("🚀 开始处理", disabled=not (data_files and match_file), use_container_width=True):
        with st.spinner("正在处理中，请稍候..."):
            processed_files, errors = process_files(data_files, match_file)
            st.session_state.processed = True
            st.session_state.processed_files = processed_files
            st.session_state.errors = errors

# 显示处理结果
if st.session_state.processed and 'processed_files' in st.session_state:
    st.markdown("---")
    
    # 统计信息
    col_stat1, col_stat2, col_stat3 = st.columns(3)
    
    with col_stat1:
        st.markdown(f"""
        <div class="stat-card">
            <div class="stat-number">{len(st.session_state.processed_files)}</div>
            <div class="stat-label">成功处理</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col_stat2:
        st.markdown(f"""
        <div class="stat-card">
            <div class="stat-number">{len(st.session_state.errors)}</div>
            <div class="stat-label">处理失败</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col_stat3:
        st.markdown(f"""
        <div class="stat-card">
            <div class="stat-number">{len(data_files)}</div>
            <div class="stat-label">总文件数</div>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown("<br>", unsafe_allow_html=True)
    
    # 下载按钮
    if st.session_state.processed_files:
        if len(st.session_state.processed_files) == 1:
            # 单个文件直接下载
            filename, content = st.session_state.processed_files[0]
            st.download_button(
                label="⬇️ 下载处理后的文件",
                data=content,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        else:
            # 多个文件打包下载
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for filename, content in st.session_state.processed_files:
                    zip_file.writestr(filename, content)
            
            st.download_button(
                label="⬇️ 下载所有处理后的文件 (ZIP)",
                data=zip_buffer.getvalue(),
                file_name="processed_files.zip",
                mime="application/zip",
                use_container_width=True
            )
    
    # 显示错误
    if st.session_state.errors:
        with st.expander("⚠️ 查看错误详情", expanded=False):
            for error in st.session_state.errors:
                st.error(error)

# 日志显示
if st.session_state.logs:
    st.markdown("---")
    st.markdown("### 📋 处理日志")
    log_container = st.container()
    with log_container:
        log_html = '<div class="log-container">'
        for log in st.session_state.logs:
            log_html += f'<div class="log-entry">{log}</div>'
        log_html += '</div>'
        st.markdown(log_html, unsafe_allow_html=True)

# 页脚
st.markdown("---")
st.markdown("""
<div style="text-align: center; color: #666; font-size: 0.85rem;">
    <p>💡 提示：程序会自动跳过损坏的文件并继续处理其他文件</p>
    <p style="color: #00a6e4;">Powered by Streamlit | © 2024</p>
</div>
""", unsafe_allow_html=True), col1))
                    
                    if not is_b0_pattern:
                        # 关键词类型判断
                        if "oneplus" in col1:
                            label = "Brand KW"  # 品牌关键词，oneplus相关短语关键词
                        else:
                            label = "Non-brand KW"  # 除了oneplus外所有关键词
                    else:
                        # ASIN 类型判断
                        if col1 in match_set:
                            label = "Brand PAT"  # OnePlus相关asin
                        else:
                            label = "CMP PAT"  # 竞手Asin(除oneplus相关asin外的所有asin)
                        
                        # 检查是否为自动广告
                        if "auto" in col2:
                            if label == "Brand PAT":
                                label = "Auto KW"  # OnePlus相关asin, 但是有标记自动广告的
                            elif label == "CMP PAT":
                                label = "Auto PAT"  # 竞手Asin, 但是有标记自动广告的
                    
                    new_ws.append([col1, col2, label])
                
                # 保存文件
                wb.save(data_file_path)
                wb.close()
                
                # 读取处理后的文件
                with open(data_file_path, 'rb') as f:
                    processed_files.append((data_file.name, f.read()))
                
                os.unlink(data_file_path)
                add_log(f"✅ 文件 {data_file.name} 处理完成")
                
            except Exception as e:
                error_msg = f"❌ 处理文件 {data_file.name} 时出错: {str(e)}"
                errors.append(error_msg)
                add_log(error_msg)
            
            # 更新进度
            progress_bar.progress((idx + 1) / total_files)
        
        status_text.text("✅ 所有文件处理完成！")
        
        return processed_files, errors
        
    except Exception as e:
        add_log(f"❌ 发生错误: {str(e)}")
        return [], [str(e)]

# 主界面
st.markdown("## 📤 文件上传")

col1, col2 = st.columns(2)

with col1:
    st.markdown("### 📁 数据文件")
    st.markdown("""
    <div style="background-color: #f0f9ff; padding: 10px; border-radius: 5px; margin-bottom: 10px;">
    <small><b>格式要求：</b></small><br>
    <small>• 文件格式：<code>.xlsx</code></small><br>
    <small>• 第1列：Query（关键词/ASIN）</small><br>
    <small>• 第5列：Campaign Type</small><br>
    <small>• 第一行为表头，数据从第二行开始</small>
    </div>
    """, unsafe_allow_html=True)
    
    data_files = st.file_uploader(
        "选择要处理的 Excel 文件（可多选）",
        type=['xlsx'],
        accept_multiple_files=True,
        key="data_files",
        help="支持同时上传多个文件进行批量处理"
    )
    
    if data_files:
        st.success(f"✅ 已选择 {len(data_files)} 个文件")
        with st.expander("📋 查看文件列表"):
            for idx, f in enumerate(data_files, 1):
                file_size = len(f.getvalue()) / 1024  # KB
                st.write(f"{idx}. 📄 {f.name} ({file_size:.1f} KB)")

with col2:
    st.markdown("### 🔍 匹配文件")
    st.markdown("""
    <div style="background-color: #fff5e6; padding: 10px; border-radius: 5px; margin-bottom: 10px;">
    <small><b>格式要求：</b></small><br>
    <small>• 文件格式：<code>.xlsx</code></small><br>
    <small>• 第1列：ASIN 列表</small><br>
    <small>• ASIN 格式：B0 开头的10位字符</small><br>
    <small>• 用于判断 Brand PAT 和 CMP PAT</small>
    </div>
    """, unsafe_allow_html=True)
    
    match_file = st.file_uploader(
        "选择包含 ASIN 的匹配文件（单个）",
        type=['xlsx'],
        key="match_file",
        help="此文件用于匹配 ASIN，判断是否为品牌产品"
    )
    
    if match_file:
        file_size = len(match_file.getvalue()) / 1024  # KB
        st.success(f"✅ 已选择: {match_file.name} ({file_size:.1f} KB)")

st.markdown("---")

# 处理按钮
col_btn1, col_btn2, col_btn3 = st.columns([1, 1, 1])
with col_btn2:
    if st.button("🚀 开始处理", disabled=not (data_files and match_file), use_container_width=True):
        with st.spinner("正在处理中，请稍候..."):
            processed_files, errors = process_files(data_files, match_file)
            st.session_state.processed = True
            st.session_state.processed_files = processed_files
            st.session_state.errors = errors

# 显示处理结果
if st.session_state.processed and 'processed_files' in st.session_state:
    st.markdown("---")
    
    # 统计信息
    col_stat1, col_stat2, col_stat3 = st.columns(3)
    
    with col_stat1:
        st.markdown(f"""
        <div class="stat-card">
            <div class="stat-number">{len(st.session_state.processed_files)}</div>
            <div class="stat-label">成功处理</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col_stat2:
        st.markdown(f"""
        <div class="stat-card">
            <div class="stat-number">{len(st.session_state.errors)}</div>
            <div class="stat-label">处理失败</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col_stat3:
        st.markdown(f"""
        <div class="stat-card">
            <div class="stat-number">{len(data_files)}</div>
            <div class="stat-label">总文件数</div>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown("<br>", unsafe_allow_html=True)
    
    # 下载按钮
    if st.session_state.processed_files:
        if len(st.session_state.processed_files) == 1:
            # 单个文件直接下载
            filename, content = st.session_state.processed_files[0]
            st.download_button(
                label="⬇️ 下载处理后的文件",
                data=content,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        else:
            # 多个文件打包下载
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for filename, content in st.session_state.processed_files:
                    zip_file.writestr(filename, content)
            
            st.download_button(
                label="⬇️ 下载所有处理后的文件 (ZIP)",
                data=zip_buffer.getvalue(),
                file_name="processed_files.zip",
                mime="application/zip",
                use_container_width=True
            )
    
    # 显示错误
    if st.session_state.errors:
        with st.expander("⚠️ 查看错误详情", expanded=False):
            for error in st.session_state.errors:
                st.error(error)

# 日志显示
if st.session_state.logs:
    st.markdown("---")
    st.markdown("### 📋 处理日志")
    log_container = st.container()
    with log_container:
        log_html = '<div class="log-container">'
        for log in st.session_state.logs:
            log_html += f'<div class="log-entry">{log}</div>'
        log_html += '</div>'
        st.markdown(log_html, unsafe_allow_html=True)

# 页脚
st.markdown("---")
st.markdown("""
<div style="text-align: center; color: #666; font-size: 0.85rem;">
    <p>💡 提示：程序会自动跳过损坏的文件并继续处理其他文件</p>
    <p style="color: #00a6e4;">Powered by Streamlit | © 2024</p>
</div>
""", unsafe_allow_html=True)
